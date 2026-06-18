# -*- coding: utf-8 -*-
"""Blog SEO Command Center — data model + import roadmap + queries + pipeline.

Quản lý 40 bài roadmap (nguồn: docs/ke_hoach_blog_seo_sintech_gsc_16m_3m.xlsx).
Bảng MỚI riêng (không đụng blog_jobs / blog_rewrite_* cũ):
  - blog_content_items   : 1 bài roadmap = 1 dòng (đủ brief + workflow + KPI)
  - blog_content_events  : audit log thao tác

Mọi migration additive + idempotent (CREATE TABLE IF NOT EXISTS / ADD COLUMN guard).
Import idempotent theo roadmap_id (re-run không nhân đôi, giữ tiến độ workflow).
"""
import json
import re
import sqlite3
import datetime
from pathlib import Path

ROOT = Path(__file__).parent
DB_PATH = ROOT / "data" / "posts.db"
XLSX_PATH = ROOT / "docs" / "ke_hoach_blog_seo_sintech_gsc_16m_3m.xlsx"

# Status workflow chuẩn
STATUSES = [
    "backlog", "brief_ready", "writing", "image_needed", "seo_review",
    "ready_publish", "published", "monitor_14d", "monitor_28d", "monitor_60d",
    "refresh_needed", "paused", "archived",
]
KANBAN_COLUMNS = [
    ("backlog", "Backlog"), ("brief_ready", "Brief Ready"), ("writing", "Writing"),
    ("image_needed", "Image Needed"), ("seo_review", "SEO Review"),
    ("ready_publish", "Ready Publish"), ("published", "Published"),
    ("monitor_14d", "Monitor"), ("refresh_needed", "Refresh Needed"),
]
# Cột workflow KHÔNG bị ghi đè khi re-import (giữ tiến độ)
_WORKFLOW_FIELDS = (
    "status", "owner", "brief_json", "draft_body_html", "draft_meta", "draft_title",
    "publish_url", "published_at", "haravan_article_id",
    "gsc_14d_json", "gsc_28d_json", "gsc_60d_json", "next_action",
)


def _now():
    return datetime.datetime.now().isoformat(timespec="seconds")


def _conn():
    c = sqlite3.connect(str(DB_PATH), timeout=30)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA busy_timeout=30000")
    return c


# ─────────────────────────── SCHEMA (additive, idempotent) ───────────────────────────
def ensure_schema():
    c = _conn()
    try:
        c.execute("""
        CREATE TABLE IF NOT EXISTS blog_content_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT DEFAULT 'roadmap',
            roadmap_id TEXT UNIQUE,
            title TEXT,
            h1 TEXT,
            title_seo TEXT,
            meta_description TEXT,
            slug TEXT,
            cluster TEXT,
            priority TEXT,
            week INTEGER,
            phase TEXT,
            intent TEXT,
            funnel TEXT,
            seo_score REAL,
            impact INTEGER,
            conversion INTEGER,
            effort INTEGER,
            status TEXT DEFAULT 'backlog',
            owner TEXT,
            deadline_draft TEXT,
            deadline_publish TEXT,
            target_url TEXT,
            cta TEXT,
            main_keyword TEXT,
            secondary_keywords_json TEXT,
            outline_json TEXT,
            eeat_assets TEXT,
            schema_type TEXT,
            internal_links_json TEXT,
            image_keywords_json TEXT,
            risk_note TEXT,
            signal_16m TEXT,
            signal_3m TEXT,
            reason TEXT,
            kpi_note TEXT,
            note TEXT,
            wordcount_target INTEGER,
            -- workflow / output
            brief_json TEXT,
            draft_title TEXT,
            draft_meta TEXT,
            draft_body_html TEXT,
            publish_url TEXT,
            published_at TEXT,
            haravan_article_id INTEGER,
            gsc_14d_json TEXT,
            gsc_28d_json TEXT,
            gsc_60d_json TEXT,
            next_action TEXT,
            import_source TEXT,
            imported_at TEXT,
            created_at TEXT,
            updated_at TEXT
        )""")
        c.execute("""
        CREATE TABLE IF NOT EXISTS blog_content_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER,
            roadmap_id TEXT,
            action TEXT,
            detail TEXT,
            at TEXT
        )""")
        c.execute("CREATE INDEX IF NOT EXISTS idx_bci_status ON blog_content_items(status)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_bci_cluster ON blog_content_items(cluster)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_bci_week ON blog_content_items(week)")
        c.commit()
    finally:
        c.close()


def log_event(item_id, roadmap_id, action, detail=""):
    c = _conn()
    try:
        c.execute("INSERT INTO blog_content_events(item_id,roadmap_id,action,detail,at) VALUES(?,?,?,?,?)",
                  (item_id, roadmap_id, action, detail[:500], _now()))
        c.commit()
    finally:
        c.close()


# ─────────────────────────── IMPORT ROADMAP (idempotent) ───────────────────────────
def _d(v):
    """Excel cell -> str/ISO date sạch."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    return str(v).strip()


def _i(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


_STATUS_MAP = {
    "chưa viết": "backlog", "chua viet": "backlog", "": "backlog", "none": "backlog",
    "đang viết": "writing", "draft": "brief_ready", "review": "seo_review",
    "đã đăng": "published", "published": "published",
}


def import_roadmap(xlsx_path=None):
    """Đọc Excel roadmap -> import/refresh vào blog_content_items. Idempotent theo roadmap_id.
    Re-run: cập nhật field KẾ HOẠCH, GIỮ field workflow (status/owner/draft/published/gsc).
    Trả report dict."""
    import openpyxl
    ensure_schema()
    path = Path(xlsx_path) if xlsx_path else XLSX_PATH
    if not path.exists():
        raise FileNotFoundError(f"Không thấy Excel roadmap: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # --- Internal links sheet -> {source_id: [ {anchor,target,purpose,priority,placement} ]} ---
    links_by_id = {}
    if "Internal Links" in wb.sheetnames:
        rows = list(wb["Internal Links"].iter_rows(values_only=True))
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            sid = str(r[0]).strip()
            links_by_id.setdefault(sid, []).append({
                "placement": _d(r[2]) if len(r) > 2 else "",
                "anchor": _d(r[3]) if len(r) > 3 else "",
                "target": _d(r[4]) if len(r) > 4 else "",
                "purpose": _d(r[5]) if len(r) > 5 else "",
                "priority": _d(r[6]) if len(r) > 6 else "",
            })

    ws = wb["Blog Roadmap"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    report = {"imported": 0, "updated": 0, "skipped": 0, "errors": 0, "ids": []}
    c = _conn()
    try:
        for r in rows[1:]:
            if not r or not r[0] or not str(r[0]).strip().upper().startswith("BLG"):
                continue
            try:
                rid = str(r[0]).strip()
                outline_raw = _d(r[22]) or ""
                outline = [h.strip() for h in outline_raw.replace("\n", "|").split("|") if h.strip()]
                secondary = [k.strip() for k in (_d(r[14]) or "").replace("\n", ",").split(",") if k.strip()]
                status_raw = (_d(r[31]) or "").lower()
                status = _STATUS_MAP.get(status_raw, "backlog")
                main_kw = _d(r[13]) or ""
                img_kws = [main_kw] if main_kw else []
                plan = {
                    "roadmap_id": rid,
                    "phase": _d(r[1]), "week": _i(r[2]),
                    "deadline_publish": _d(r[34]) or _d(r[3]),
                    "cluster": _d(r[4]), "priority": _d(r[5]),
                    "intent": _d(r[6]), "funnel": _d(r[7]),
                    "title": _d(r[8]), "h1": _d(r[8]),
                    "title_seo": _d(r[9]), "meta_description": _d(r[11]),
                    "main_keyword": main_kw,
                    "secondary_keywords_json": json.dumps(secondary, ensure_ascii=False),
                    "signal_16m": _d(r[15]), "signal_3m": _d(r[16]), "reason": _d(r[17]),
                    "slug": _d(r[18]), "target_url": _d(r[19]),
                    "cta": _d(r[21]),
                    "outline_json": json.dumps(outline, ensure_ascii=False),
                    "eeat_assets": _d(r[23]), "schema_type": _d(r[24]),
                    "wordcount_target": _i(r[25]),
                    "impact": _i(r[26]), "conversion": _i(r[27]), "effort": _i(r[28]),
                    "risk_note": _d(r[29]), "seo_score": _f(r[30]),
                    "owner": _d(r[32]), "deadline_draft": _d(r[33]),
                    "kpi_note": _d(r[35]), "note": _d(r[36]),
                    "internal_links_json": json.dumps(links_by_id.get(rid, []), ensure_ascii=False),
                    "image_keywords_json": json.dumps(img_kws, ensure_ascii=False),
                    "import_source": path.name, "imported_at": _now(), "updated_at": _now(),
                }
                existing = c.execute("SELECT id,owner FROM blog_content_items WHERE roadmap_id=?", (rid,)).fetchone()
                if existing:
                    # update KẾ HOẠCH fields, GIỮ workflow; owner: chỉ set nếu chưa có
                    fields = dict(plan)
                    if existing["owner"]:
                        fields.pop("owner", None)
                    # GIỮ title_seo + meta_description đã tinh chỉnh trên web (không để Excel ghi đè)
                    fields.pop("title_seo", None)
                    fields.pop("meta_description", None)
                    cols = ", ".join(f"{k}=?" for k in fields)
                    c.execute(f"UPDATE blog_content_items SET {cols} WHERE roadmap_id=?",
                              (*fields.values(), rid))
                    report["updated"] += 1
                else:
                    plan["source"] = "roadmap"
                    plan["status"] = status
                    plan["created_at"] = _now()
                    cols = ", ".join(plan.keys())
                    ph = ", ".join("?" for _ in plan)
                    c.execute(f"INSERT INTO blog_content_items({cols}) VALUES({ph})", tuple(plan.values()))
                    report["imported"] += 1
                report["ids"].append(rid)
            except Exception as e:
                report["errors"] += 1
                report.setdefault("error_detail", []).append(str(e)[:120])
        c.commit()
    finally:
        c.close()
    report["total_in_db"] = count_items()
    return report


# ─────────────────────────── QUERIES ───────────────────────────
def count_items():
    c = _conn()
    try:
        return c.execute("SELECT COUNT(*) FROM blog_content_items WHERE source='roadmap'").fetchone()[0]
    finally:
        c.close()


def _row_to_dict(row):
    return {k: row[k] for k in row.keys()}


def list_items(cluster=None, priority=None, status=None, week=None, owner_missing=False,
               overdue=False, q=None):
    c = _conn()
    try:
        sql = "SELECT * FROM blog_content_items WHERE source='roadmap'"
        args = []
        if cluster:
            sql += " AND cluster=?"; args.append(cluster)
        if priority:
            sql += " AND priority=?"; args.append(priority)
        if status:
            sql += " AND status=?"; args.append(status)
        if week:
            sql += " AND week=?"; args.append(int(week))
        if owner_missing:
            sql += " AND (owner IS NULL OR owner='')"
        if q:
            sql += " AND (title LIKE ? OR main_keyword LIKE ?)"; args += [f"%{q}%", f"%{q}%"]
        sql += " ORDER BY week, priority, roadmap_id"
        rows = [_row_to_dict(r) for r in c.execute(sql, args).fetchall()]
    finally:
        c.close()
    today = datetime.date.today().isoformat()
    for r in rows:
        r["is_overdue"] = bool(r.get("deadline_publish") and r["deadline_publish"] < today
                               and r.get("status") != "published")
    if overdue:
        rows = [r for r in rows if r["is_overdue"]]
    return rows


def get_item(item_id):
    c = _conn()
    try:
        row = c.execute("SELECT * FROM blog_content_items WHERE id=?", (item_id,)).fetchone()
        return _row_to_dict(row) if row else None
    finally:
        c.close()


def status_summary():
    items = list_items()
    today = datetime.date.today().isoformat()
    by_status = {}
    for it in items:
        by_status[it["status"]] = by_status.get(it["status"], 0) + 1
    return {
        "total": len(items),
        "a1": sum(1 for i in items if i["priority"] == "A1"),
        "a2": sum(1 for i in items if i["priority"] == "A2"),
        "writing": by_status.get("writing", 0),
        "review": by_status.get("seo_review", 0),
        "ready_publish": by_status.get("ready_publish", 0),
        "published": by_status.get("published", 0),
        "overdue": sum(1 for i in items if i["is_overdue"]),
        "no_owner": sum(1 for i in items if not i.get("owner")),
        "by_status": by_status,
        "by_cluster": _group_count(items, "cluster"),
    }


def _group_count(items, key):
    out = {}
    for i in items:
        out[i.get(key) or "(khác)"] = out.get(i.get(key) or "(khác)", 0) + 1
    return out


def kanban():
    items = list_items()
    cols = []
    for key, label in KANBAN_COLUMNS:
        if key == "monitor_14d":
            group = [i for i in items if i["status"] in ("monitor_14d", "monitor_28d", "monitor_60d")]
        else:
            group = [i for i in items if i["status"] == key]
        cols.append({"key": key, "label": label, "count": len(group),
                     "items": [{"id": i["id"], "roadmap_id": i["roadmap_id"], "title": i["title"],
                                "priority": i["priority"], "cluster": i["cluster"],
                                "is_overdue": i["is_overdue"]} for i in group]})
    warnings = []
    wmap = {c["key"]: c["count"] for c in cols}
    if wmap.get("writing", 0) > 5:
        warnings.append(f"WIP cao: Writing = {wmap['writing']} (>5)")
    if wmap.get("seo_review", 0) > 5:
        warnings.append(f"WIP cao: SEO Review = {wmap['seo_review']} (>5)")
    return {"columns": cols, "warnings": warnings}


def calendar():
    items = list_items()
    today = datetime.date.today().isoformat()
    weeks = {}
    for i in items:
        w = i.get("week") or 0
        weeks.setdefault(w, {"week": w, "items": []})
        weeks[w]["items"].append({
            "id": i["id"], "roadmap_id": i["roadmap_id"], "title": i["title"],
            "cluster": i["cluster"], "priority": i["priority"], "status": i["status"],
            "deadline_draft": i.get("deadline_draft"), "deadline_publish": i.get("deadline_publish"),
            "is_overdue": i["is_overdue"], "no_owner": not i.get("owner"),
            "no_brief": not i.get("brief_json"),
        })
    return {"weeks": [weeks[k] for k in sorted(weeks)]}


def kpi_monitor():
    """KPI sau publish (đọc cột gsc_* đã lưu — KHÔNG gọi API khi render)."""
    items = list_items(status="published") + [i for i in list_items() if i.get("published_at")]
    seen = set(); out = []
    today = datetime.date.today()
    for i in items:
        if i["id"] in seen:
            continue
        seen.add(i["id"])
        pub = i.get("published_at")
        milestones = {}
        if pub:
            try:
                pub_d = datetime.date.fromisoformat(pub[:10])
                for d, key in ((14, "14d"), (28, "28d"), (60, "60d")):
                    due = pub_d + datetime.timedelta(days=d)
                    has = bool(i.get(f"gsc_{key}_json"))
                    milestones[key] = {"due": due.isoformat(), "reached": today >= due, "has_data": has}
            except ValueError:
                pass
        out.append({"id": i["id"], "roadmap_id": i["roadmap_id"], "title": i["title"],
                    "publish_url": i.get("publish_url"), "published_at": pub,
                    "milestones": milestones, "next_action": i.get("next_action")})
    return {"items": out, "note": "Số liệu đọc từ DB/cache; chưa tracking thì hiện 'chưa có data'."}


# ─────────────────────────── MUTATIONS ───────────────────────────
def set_field(item_id, field, value, allowed):
    if field not in allowed:
        raise ValueError(f"field không cho phép: {field}")
    it = get_item(item_id)
    if not it:
        raise ValueError("item không tồn tại")
    c = _conn()
    try:
        c.execute(f"UPDATE blog_content_items SET {field}=?, updated_at=? WHERE id=?",
                  (value, _now(), item_id))
        c.commit()
    finally:
        c.close()
    log_event(item_id, it["roadmap_id"], f"set_{field}", str(value)[:120])
    return get_item(item_id)


def set_status(item_id, status):
    if status not in STATUSES:
        raise ValueError(f"status không hợp lệ: {status}")
    return set_field(item_id, "status", status, {"status"})


# ─────────────────────────── PIPELINE (brief / draft / sync) ───────────────────────────
def _ddmm(iso):
    """ISO 'YYYY-MM-DD' -> 'DD/MM' (cho UI export3)."""
    if not iso:
        return ""
    try:
        d = datetime.date.fromisoformat(iso[:10])
        return f"{d.day:02d}/{d.month:02d}"
    except ValueError:
        return iso


_CLUSTER_TONE = {"Build PC": "violet", "Linh kiện cũ": "sky", "Phần mềm bản quyền": "teal",
                 "Dịch vụ Quận 7": "amber", "Kỹ thuật/Gaming hỗ trợ": "indigo"}
_PRI_TONE = {"A1": "rose", "A2": "amber", "A3": "sky"}
_STATUS_META = {
    "backlog": {"label": "Backlog", "tone": "slate", "col": "Backlog"},
    "brief_ready": {"label": "Brief ready", "tone": "sky", "col": "Brief Ready"},
    "writing": {"label": "Đang viết", "tone": "violet", "col": "Writing"},
    "image_needed": {"label": "Cần ảnh", "tone": "amber", "col": "Image Needed"},
    "seo_review": {"label": "SEO review", "tone": "indigo", "col": "SEO Review"},
    "ready_publish": {"label": "Sẵn sàng đăng", "tone": "teal", "col": "Ready Publish"},
    "published": {"label": "Đã đăng", "tone": "teal", "col": "Published"},
    "monitor_14d": {"label": "Monitor 14d", "tone": "sky", "col": "Monitor"},
    "monitor_28d": {"label": "Monitor 28d", "tone": "sky", "col": "Monitor"},
    "monitor_60d": {"label": "Monitor 60d", "tone": "sky", "col": "Monitor"},
    "refresh_needed": {"label": "Cần refresh", "tone": "rose", "col": "Refresh Needed"},
    "paused": {"label": "Tạm dừng", "tone": "slate", "col": "Backlog"},
    "archived": {"label": "Lưu trữ", "tone": "slate", "col": "Backlog"},
}
_KANBAN_COLS = ["Backlog", "Brief Ready", "Writing", "Image Needed", "SEO Review",
                "Ready Publish", "Published", "Monitor", "Refresh Needed"]


def _funnel_tone(funnel):
    f = (funnel or "").upper()
    if "BOFU" in f:
        return "teal"
    if "MOFU" in f:
        return "violet"
    return "sky"


def _kpi_from_json(raw):
    if not raw:
        return None
    try:
        d = json.loads(raw)
        return {"clicks": d.get("clicks", 0), "impr": d.get("impr", d.get("impressions", 0)),
                "ctr": d.get("ctr", 0), "pos": d.get("pos", d.get("position", 0))}
    except (ValueError, TypeError):
        return None


def _item_to_ui(it):
    sec = json.loads(it.get("secondary_keywords_json") or "[]")
    outline = json.loads(it.get("outline_json") or "[]")
    links = json.loads(it.get("internal_links_json") or "[]")
    internal = [(l.get("target") or l.get("anchor") or "") if isinstance(l, dict) else str(l) for l in links]
    images = json.loads(it.get("image_keywords_json") or "[]")
    eeat = [s.strip() for s in re.split(r"[·;,\n]", it.get("eeat_assets") or "") if s.strip()]
    return {
        "id": it["roadmap_id"], "pri": it.get("priority"), "week": it.get("week"),
        "phase": (it.get("phase") or "").replace("GĐ", "").strip()[:1] or it.get("week"),
        "cluster": it.get("cluster"), "title": it.get("title"),
        "intent": it.get("intent"), "funnel": it.get("funnel"), "seo": it.get("seo_score") or 0,
        "impact": it.get("impact") or 0, "status": it.get("status"), "owner": it.get("owner") or "",
        "dDraft": _ddmm(it.get("deadline_draft")), "dPub": _ddmm(it.get("deadline_publish")),
        "kw": it.get("main_keyword") or "", "next": it.get("next_action") or "",
        "overdue": bool(it.get("is_overdue")), "slug": it.get("slug") or "",
        "titleSeo": it.get("title_seo") or it.get("title"), "meta": it.get("meta_description") or "",
        "h1": it.get("h1") or it.get("title"), "secondary": sec,
        "target": it.get("target_url") or "", "cta": it.get("cta") or "",
        "outline": outline, "eeat": eeat, "schema": it.get("schema_type") or "",
        "internal": internal, "images": images, "risk": it.get("risk_note") or "",
        "kpi": {"d14": _kpi_from_json(it.get("gsc_14d_json")),
                "d28": _kpi_from_json(it.get("gsc_28d_json")),
                "d60": _kpi_from_json(it.get("gsc_60d_json"))},
        "dbid": it["id"],
        "audit": _audit_for(it),
    }


def _audit_for(it):
    log = []
    imp = _ddmm(it.get("imported_at"))
    log.append({"t": imp or "—", "a": "Tạo từ roadmap import", "who": "system"})
    if it.get("status") and it["status"] != "backlog":
        log.append({"t": imp or "—", "a": "Brief sẵn sàng / đang xử lý", "who": it.get("owner") or "—"})
    if it.get("published_at"):
        log.append({"t": _ddmm(it.get("published_at")), "a": "Đăng lên Haravan", "who": it.get("owner") or "—"})
    return log


def _nav():
    """Nav sidebar với link THẬT của app (cho shell export3)."""
    return [
        {"group": "Tổng quan", "items": [
            {"icon": "layout-dashboard", "label": "Dashboard", "url": "/"},
            {"icon": "list-checks", "label": "Job Center", "url": "/jobs"},
        ]},
        {"group": "Phân tích & đo lường", "items": [
            {"icon": "chart-area", "label": "GA4 Analytics", "url": "/seo/ga4"},
            {"icon": "search", "label": "Search Console", "url": "/seo/gsc"},
            {"icon": "activity", "label": "Analytics Ops", "url": "/ops/analytics"},
        ]},
        {"group": "Facebook", "items": [
            {"icon": "plus-circle", "label": "Bài mới", "url": "/posts/new"},
            {"icon": "send", "label": "Bài đã đăng", "url": "/posts"},
            {"icon": "calendar-days", "label": "Lịch đăng", "url": "/calendar"},
        ]},
        {"group": "Nội dung", "items": [
            {"icon": "package", "label": "Content Jobs", "url": "/content-jobs"},
            {"icon": "folder-tree", "label": "Collection Content", "url": "/collection-content"},
            {"icon": "file-text", "label": "Blog Command Center", "url": "/blog-content", "active": True},
            {"icon": "shopping-bag", "label": "SP mới", "url": "/products/new"},
            {"icon": "image", "label": "ALT ảnh", "url": "/alt-manager"},
        ]},
        {"group": "Tối ưu & kho", "items": [
            {"icon": "shield-check", "label": "SEO Title/Meta", "url": "/seo/title-meta"},
            {"icon": "search", "label": "SEO Dashboard", "url": "/seo"},
        ]},
    ]


def ui_payload():
    """Payload cho UI Command Center (export3 shell). Inject window.SHB + window.SH."""
    items = [_item_to_ui(it) for it in list_items()]
    shb = {
        "items": items, "clusterTone": _CLUSTER_TONE, "priTone": _PRI_TONE,
        "funnelTone": {f: _funnel_tone(f) for f in set(i["funnel"] for i in items if i["funnel"])},
        "statusMeta": _STATUS_META, "clusters": list(_CLUSTER_TONE.keys()), "kanbanCols": _KANBAN_COLS,
    }
    return {"SHB": shb, "SH": {"nav": _nav()}}


def save_edit(item_id, title=None, meta=None, h1=None, slug=None, body=None):
    """Lưu nội dung vợ tự soạn trong editor (title/meta/h1/slug + body_html draft).
    KHÔNG đụng title/meta gốc roadmap (giữ tham khảo) — ghi vào draft_* + title_seo/meta hiện hành."""
    it = get_item(item_id)
    if not it:
        raise ValueError("item không tồn tại")
    sets, args = [], []
    if title is not None:
        sets.append("title_seo=?"); args.append(title)
    if meta is not None:
        sets.append("meta_description=?"); args.append(meta)
    if h1 is not None:
        sets.append("h1=?"); args.append(h1)
    if slug is not None:
        sets.append("slug=?"); args.append(slug)
    if body is not None:
        sets += ["draft_body_html=?", "draft_title=?", "draft_meta=?"]
        args += [body, title if title is not None else it.get("title_seo"),
                 meta if meta is not None else it.get("meta_description")]
        # nếu đang backlog/brief_ready và đã có body → chuyển sang writing
        if it.get("status") in ("backlog", "brief_ready") and body.strip():
            sets.append("status='writing'")
    if not sets:
        return get_item(item_id)
    sets.append("updated_at=?"); args.append(_now())
    args.append(item_id)
    c = _conn()
    try:
        c.execute(f"UPDATE blog_content_items SET {', '.join(sets)} WHERE id=?", args)
        c.commit()
    finally:
        c.close()
    log_event(item_id, it["roadmap_id"], "save_edit",
              f"title={'y' if title is not None else '-'} body={len(body or '')}c")
    return get_item(item_id)


def build_brief(item_id):
    """Tạo brief từ dữ liệu roadmap (KHÔNG gọi AI) -> lưu brief_json, status->brief_ready."""
    it = get_item(item_id)
    if not it:
        raise ValueError("item không tồn tại")
    brief = {
        "roadmap_id": it["roadmap_id"], "title": it["title_seo"] or it["title"],
        "h1": it["h1"], "meta": it["meta_description"], "slug": it["slug"],
        "main_keyword": it["main_keyword"],
        "secondary_keywords": json.loads(it.get("secondary_keywords_json") or "[]"),
        "intent": it["intent"], "funnel": it["funnel"],
        "outline": json.loads(it.get("outline_json") or "[]"),
        "target_url": it["target_url"], "cta": it["cta"],
        "internal_links": json.loads(it.get("internal_links_json") or "[]"),
        "eeat_assets": it["eeat_assets"], "schema": it["schema_type"],
        "wordcount_target": it["wordcount_target"], "risk_note": it["risk_note"],
        "rule": "KHÔNG viết hướng dẫn crack/kích hoạt lậu; gắn internal link + CTA Sintech; không bịa thông số.",
    }
    c = _conn()
    try:
        c.execute("UPDATE blog_content_items SET brief_json=?, status=?, updated_at=? WHERE id=?",
                  (json.dumps(brief, ensure_ascii=False),
                   "brief_ready" if it["status"] == "backlog" else it["status"], _now(), item_id))
        c.commit()
    finally:
        c.close()
    log_event(item_id, it["roadmap_id"], "generate_brief", "brief built from roadmap")
    return brief


def generate_draft(item_id):
    """Gọi blog_content_writer theo brief roadmap -> lưu draft local (KHÔNG publish)."""
    import blog_content_writer as bcw
    it = get_item(item_id)
    if not it:
        raise ValueError("item không tồn tại")
    brief = json.loads(it["brief_json"]) if it.get("brief_json") else build_brief(item_id)
    # Map sang brief mà gen_blog_content_from_brief mong đợi
    gen_brief = {
        "article_title": brief.get("title") or it["title"],
        "keyword": brief.get("main_keyword", ""),
        "intent": brief.get("intent", ""),
        "unique_angle": (it.get("reason") or "")[:200],
        "outline": "\n".join(brief.get("outline", [])),
        "target_blog": "huong-dan" if "huong" in (it.get("target_url") or "") else "news",
    }
    res = bcw.gen_blog_content_from_brief(gen_brief)
    body = res.get("body_html") or res.get("content") or ""
    title = res.get("title") or brief.get("title")
    meta = res.get("meta") or res.get("meta_description") or brief.get("meta")
    c = _conn()
    try:
        c.execute("""UPDATE blog_content_items SET draft_title=?, draft_meta=?, draft_body_html=?,
                     status=CASE WHEN status IN ('backlog','brief_ready') THEN 'writing' ELSE status END,
                     updated_at=? WHERE id=?""",
                  (title, meta, body, _now(), item_id))
        c.commit()
    finally:
        c.close()
    log_event(item_id, it["roadmap_id"], "generate_draft", f"draft {len(body)} chars")
    return {"ok": bool(body), "title": title, "meta": meta, "body_len": len(body)}


SYNC_CONFIRM_PHRASE = "PUBLISH BLOG ITEM"


def sync_item(item_id, confirm_phrase, publish=False):
    """Đẩy bài lên Haravan — flow đầy đủ (title/meta/SEO page_title/author/handle/tags/body),
    KHÔNG kèm ảnh (ảnh tự set trong admin). GATED bằng confirm phrase.
    publish=False (mặc định) -> bài ẨN (chưa đăng, không lên Google). True -> đăng công khai.
    Có aid -> UPDATE bài cũ (backup body trước, xử lý 500-but-write); chưa có -> CREATE ẩn."""
    if confirm_phrase != SYNC_CONFIRM_PHRASE:
        return {"ok": False, "error": "GATE: cần confirm phrase đúng để sync."}
    import haravan_blog as hb
    it = get_item(item_id)
    if not it:
        return {"ok": False, "error": "item không tồn tại"}
    body = it.get("draft_body_html") or ""
    if not body.strip():
        return {"ok": False, "error": "Chưa có nội dung bài — hãy soạn + Lưu bài trước khi sync."}
    slug = (it.get("slug") or "").rstrip("/")
    blog_id = 1000960873 if "huong" in slug else 1000906526   # huong-dan vs news
    handle = slug.split("/")[-1] or None
    sec = json.loads(it.get("secondary_keywords_json") or "[]")
    tags = ", ".join([it.get("main_keyword") or ""] + sec[:3]).strip(", ")
    title = it.get("title_seo") or it.get("h1") or it.get("title")
    meta = it.get("meta_description") or ""
    fields = {"title": title, "body_html": body, "meta_description": meta,
              "page_title": title, "summary_html": meta, "author": "Trọng Nghĩa"}
    if handle:
        fields["handle"] = handle
    if tags:
        fields["tags"] = tags
    bdir = ROOT / "data" / "_blog_content_sync_backup"
    bdir.mkdir(parents=True, exist_ok=True)
    aid = it.get("haravan_article_id")
    try:
        if aid:
            try:
                live = hb.get_article(blog_id, aid)
                (bdir / f"item_{item_id}_article_{aid}.html").write_text(
                    live.get("body_html") or "", encoding="utf-8")
            except Exception:
                pass
            try:
                hb.update_article(blog_id, aid, fields)
            except Exception as e:
                # 500-but-write: verify body đã ghi mới chưa
                if "500" in str(e):
                    chk = hb.get_article(blog_id, aid)
                    if (chk.get("body_html") or "")[:160] != body[:160]:
                        raise
                else:
                    raise
            new_aid, action = aid, "update"
        else:
            art = hb.create_article(blog_id, fields, hidden=not publish)
            new_aid, action = art.get("id"), "create"
        if publish and new_aid:
            hb.update_article(blog_id, new_aid, {"published": True})
        c = _conn()
        try:
            c.execute("""UPDATE blog_content_items SET haravan_article_id=?, publish_url=?,
                         published_at=CASE WHEN ? THEN ? ELSE published_at END,
                         status=CASE WHEN ? THEN 'published' ELSE status END, updated_at=? WHERE id=?""",
                      (new_aid, handle, publish, _now(), publish, _now(), item_id))
            c.commit()
        finally:
            c.close()
        log_event(item_id, it["roadmap_id"], "sync", f"{action} blog={blog_id} aid={new_aid} hidden={not publish}")
        return {"ok": True, "article_id": new_aid, "action": action, "hidden": not publish,
                "admin": f"https://sintech.myharavan.com/admin/sale_channels/online_store/blogs/{blog_id}/articles/{new_aid}"}
    except Exception as e:
        log_event(item_id, it["roadmap_id"], "sync_error", str(e)[:200])
        return {"ok": False, "error": str(e)[:250]}
