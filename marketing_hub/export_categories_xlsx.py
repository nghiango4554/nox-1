import urllib.request
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import local_config as _lcfg

SITEMAP_URL = "https://sintech.vn/sitemap_collections_1.xml"
_desktop_sintech = _lcfg.get("DESKTOP_SINTECH", r"C:\Users\NGHIANGO\Desktop\Sintech")
OUT_DIR = os.path.join(_desktop_sintech, "Chồng iu", "Tài liệu")
OUT_FILE = os.path.join(OUT_DIR, "Sintech_Categories.xlsx")

NAME_MAP = {
    "pc-gaming-do-hoa-ai": "PC Gaming Đồ họa AI",
    "pc-gaming": "PC Gaming",
    "pc-gaming-cao-cap": "PC Gaming Cao Cấp",
    "pc-streaming-livestream": "PC Streaming / Livestream",
    "pc-esport": "PC Esport",
    "pc-gaming-theo-gia": "PC Gaming theo Giá",
    "pc-gaming-10-20-trieu": "PC Gaming 10-20 triệu",
    "pc-gaming-20-30-trieu": "PC Gaming 20-30 triệu",
    "pc-gaming-30-50-trieu": "PC Gaming 30-50 triệu",
    "pc-gaming-50-80-trieu": "PC Gaming 50-80 triệu",
    "pc-gaming-80-100-trieu": "PC Gaming 80-100 triệu",
    "pc-gaming-tren-100-trieu": "PC Gaming trên 100 triệu",
    "pc-gaming-theo-vga": "PC Gaming theo VGA",
    "pc-rtx-3050-4050": "PC RTX 3050/4050",
    "pc-rtx-4060-5060": "PC RTX 4060/5060",
    "pc-rtx-4070-5070": "PC RTX 4070/5070",
    "pc-rtx-4080-5080": "PC RTX 4080/5080",
    "pc-rtx-4090-5090": "PC RTX 4090/5090",
    "pc-do-hoa": "PC Đồ họa",
    "pc-photoshop": "PC Photoshop",
    "pc-video-editing": "PC Video Editing",
    "pc-3d-rendering": "PC 3D Rendering",
    "pc-autocad": "PC AutoCAD",
    "pc-ai-workstation": "PC AI Workstation",
    "pc-van-phong": "PC Văn Phòng",
    "pc-van-phong-gia-re": "PC Văn Phòng Giá Rẻ",
    "pc-van-phong-doanh-nghiep": "PC Văn Phòng Doanh Nghiệp",
    "pc-hoc-tap-online": "PC Học Tập Online",
    "pc-ke-toan": "PC Kế Toán",
    "pc-mini-pc": "Mini PC",
    "pc-van-phong-may-bo": "PC Văn Phòng Máy Bộ",
    "pc-van-phong-theo-gia": "PC Văn Phòng theo Giá",
    "pc-duoi-7-trieu": "PC dưới 7 triệu",
    "pc-7-10-trieu": "PC 7-10 triệu",
    "pc-10-15-trieu": "PC 10-15 triệu",
    "pc-15-20-trieu": "PC 15-20 triệu",
    "pc-tren-20-trieu": "PC trên 20 triệu",
    "may-tinh-bo": "Máy tính bộ",
    "may-bo-dell": "Máy bộ Dell",
    "may-bo-hp": "Máy bộ HP",
    "may-bo-lenovo": "Máy bộ Lenovo",
    "may-bo-rosa": "Máy bộ Rosa",
    "all-in-one": "All in One",
    "laptop": "Laptop",
    "laptop-gaming": "Laptop Gaming",
    "laptop-van-phong": "Laptop Văn Phòng",
    "laptop-do-hoa": "Laptop Đồ họa",
    "laptop-hoc-tap-sinh-vien": "Laptop Học Tập / Sinh Viên",
    "laptop-theo-gia": "Laptop theo Giá",
    "laptop-duoi-15-trieu": "Laptop dưới 15 triệu",
    "laptop-15-20-trieu": "Laptop 15-20 triệu",
    "laptop-20-25-trieu": "Laptop 20-25 triệu",
    "laptop-25-30-trieu": "Laptop 25-30 triệu",
    "laptop-30-40-trieu": "Laptop 30-40 triệu",
    "laptop-theo-hang": "Laptop theo Hãng",
    "laptop-asus": "Laptop Asus",
    "laptop-acer": "Laptop Acer",
    "laptop-msi": "Laptop MSI",
    "laptop-lenovo": "Laptop Lenovo",
    "laptop-dell": "Laptop Dell",
    "laptop-hp": "Laptop HP",
    "laptop-macbook": "Laptop & Macbook",
    "macbook": "Macbook",
    "macbook-air": "Macbook Air",
    "macbook-pro": "Macbook Pro",
    "sac-macbook": "Sạc Macbook",
    "hang-cu": "Hàng cũ",
    "man-hinh-may-tinh": "Màn hình máy tính",
    "man-hinh-may-tinh-pc": "Màn hình máy tính PC",
    "man-hinh-gaming": "Màn hình Gaming",
    "man-hinh-do-hoa": "Màn hình Đồ họa",
    "man-hinh-van-phong": "Màn hình Văn Phòng",
    "man-hinh-cong": "Màn hình Cong",
    "man-hinh-msi": "Màn hình MSI",
    "man-hinh-vsp": "Màn hình VSP",
    "man-hinh-dell": "Màn hình Dell",
    "man-hinh-asus": "Màn hình Asus",
    "man-hinh-viewsonic": "Màn hình ViewSonic",
    "man-hinh-samsung": "Màn hình Samsung",
    "man-hinh-acer": "Màn hình Acer",
    "man-hinh-aoc": "Màn hình AOC",
    "man-hinh-aiwa": "Màn hình Aiwa",
    "man-hinh-ktc": "Màn hình KTC",
    "man-hinh-22-25-inch": "Màn hình 22-25 inch",
    "man-hinh-27-29-inch": "Màn hình 27-29 inch",
    "man-hinh-30-32-inch": "Màn hình 30-32 inch",
    "man-hinh-34-35-inch": "Màn hình 34-35 inch",
    "man-hinh-tren-43-inch": "Màn hình trên 43 inch",
    "man-hinh-full-hd": "Màn hình Full HD",
    "man-hinh-2k-qhd": "Màn hình 2K / QHD",
    "man-hinh-4k": "Màn hình 4K",
    "man-hinh-75hz-tro-xuong": "Màn hình 75Hz trở xuống",
    "man-hinh-100hz-144hz": "Màn hình 100Hz - 144Hz",
    "man-hinh-165hz-180hz": "Màn hình 165Hz - 180Hz",
    "man-hinh-240hz-tro-len": "Màn hình 240Hz trở lên",
    "man-hinh-laptop": "Màn hình Laptop",
    "gia-treo-man-hinh": "Giá treo màn hình",
    "gia-treo-don": "Giá treo đơn",
    "gia-treo-doi": "Giá treo đôi",
    "linh-kien-may-tinh-1": "Linh kiện máy tính",
    "phu-kien-may-tinh": "Phụ kiện máy tính",
    "cpu": "CPU",
    "cpu-intel": "CPU Intel",
    "cpu-amd": "CPU AMD",
    "vga": "VGA",
    "vga-nvidia": "VGA NVIDIA",
    "vga-amd": "VGA AMD",
    "mainboard": "Mainboard",
    "mainboard-intel-1": "Mainboard Intel",
    "mainboard-amd": "Mainboard AMD",
    "ram": "RAM",
    "ram-ddr4": "RAM DDR4",
    "ram-ddr5": "RAM DDR5",
    "ram-laptop": "RAM Laptop",
    "ssd-hdd-1": "SSD / HDD",
    "ssd-m-2": "SSD M.2",
    "ssd-sata": "SSD SATA",
    "ssd-laptop": "SSD Laptop",
    "hdd": "HDD",
    "psu-nguon": "PSU / Nguồn",
    "nguon-duoi-550w": "Nguồn dưới 550W",
    "nguon-600w-750w": "Nguồn 600W - 750W",
    "nguon-850w-tro-len": "Nguồn 850W trở lên",
    "case": "Case",
    "case-gaming-1": "Case Gaming",
    "case-van-phong-1": "Case Văn Phòng",
    "tan-nhiet": "Tản nhiệt",
    "tan-khi": "Tản khí",
    "tan-aio": "Tản AIO",
    "fan-case": "Fan Case",
    "fan-case-1": "Fan Case (mới)",
    "ban-phim": "Bàn phím",
    "ban-phim-gaming": "Bàn phím Gaming",
    "ban-phim-co": "Bàn phím Cơ",
    "ban-phim-khong-day-1": "Bàn phím không dây",
    "ban-phim-co-day": "Bàn phím có dây",
    "ban-phim-laptop": "Bàn phím Laptop",
    "chuot-may-tinh": "Chuột máy tính",
    "chuot-gaming": "Chuột Gaming",
    "chuot-van-phong": "Chuột Văn Phòng",
    "chuot-khong-day": "Chuột không dây",
    "chuot-co-day": "Chuột có dây",
    "ban-di-chuot": "Bàn di chuột",
    "ban-di-chuot-nho": "Bàn di chuột nhỏ",
    "ban-di-chuot-lon": "Bàn di chuột lớn",
    "tai-nghe": "Tai nghe",
    "tai-nghe-gaming": "Tai nghe Gaming",
    "tai-nghe-khong-day": "Tai nghe không dây",
    "tai-nghe-co-day-1": "Tai nghe có dây",
    "loa-1": "Loa",
    "ban-2": "Bàn",
    "ban-gaming-chu-z": "Bàn Gaming chữ Z",
    "ban-gaming-chu-k": "Bàn Gaming chữ K",
    "ban-gaming-chu-yn": "Bàn Gaming chữ YN",
    "ghe": "Ghế",
    "ghe-van-phong-2": "Ghế văn phòng",
    "ban-ghe-gaming": "Bàn ghế Gaming",
    "phu-kien-thiet-bi-mang-1": "Phụ kiện thiết bị mạng",
    "thiet-bi-mang-1": "Thiết bị mạng",
    "thiet-bi-wifi": "Thiết bị Wifi",
    "wifi-mesh": "Wifi Mesh",
    "router-wifi-1": "Router Wifi",
    "bo-phat-wifi-4g-lte": "Bộ phát Wifi 4G LTE",
    "thiet-bi-thu-wifi-bluetooth": "Thiết bị thu Wifi / Bluetooth",
    "usb-wifi-1": "USB Wifi",
    "usb-bluetooth": "USB Bluetooth",
    "bo-chia-mang-day-mang": "Bộ chia mạng / Dây mạng",
    "bo-chia-mang": "Bộ chia mạng",
    "day-mang": "Dây mạng",
    "cap-chuyen-doi": "Cáp chuyển đổi",
    "cap-hdmi": "Cáp HDMI",
    "cap-displayport": "Cáp DisplayPort",
    "cap-type-c": "Cáp Type-C",
    "day-nguon": "Dây nguồn",
    "cap-nguon-pc": "Cáp nguồn PC",
    "cap-nguon-laptop": "Cáp nguồn Laptop",
    "hub": "Hub",
    "hub-usb": "Hub USB",
    "hub-type-c": "Hub Type-C",
    "luu-tru-va-box": "Lưu trữ & Box",
    "the-nho": "Thẻ nhớ",
    "usb-flash": "USB Flash",
    "box-ssd-m-2": "Box SSD M.2",
    "box-ssd-sata": "Box SSD SATA",
    "box-hdd": "Box HDD",
    "camera": "Camera",
    "camera-trong-nha": "Camera trong nhà",
    "camera-ngoai-troi": "Camera ngoài trời",
    "may-in": "Máy in",
    "may-in-cam-day": "Máy in cắm dây",
    "may-in-wifi": "Máy in Wifi",
    "gaming-gear": "Gaming Gear",
    "phu-kien-linh-kien-laptop": "Phụ kiện / Linh kiện Laptop",
    "sac-laptop": "Sạc Laptop",
    "pin-laptop": "Pin Laptop",
    "de-tan-nhiet": "Đế tản nhiệt",
    "balo-tui-laptop": "Balo / Túi Laptop",
    "keo-thom-laptop": "Keo thơm Laptop",
    "dich-vu-ve-sinh-pc": "Dịch vụ vệ sinh PC",
    "ve-sinh-pc-tai-cua-hang": "Vệ sinh PC tại cửa hàng",
    "ve-sinh-laptop": "Vệ sinh Laptop",
    "ve-sinh-pc-tan-nha": "Vệ sinh PC tận nhà",
    "sua-chua-may-tinh": "Sửa chữa máy tính",
    "sua-chua-pc-laptop": "Sửa chữa PC / Laptop",
    "dich-vu-sua-chua": "Dịch vụ sửa chữa",
    "nang-cap-pc": "Nâng cấp PC",
    "cai-dat-windows-phan-mem": "Cài đặt Windows / Phần mềm",
}

def slug_to_name(slug):
    if slug in NAME_MAP:
        return NAME_MAP[slug]
    return slug.replace("-", " ").title()

req = urllib.request.Request(SITEMAP_URL, headers={"User-Agent": "Mozilla/5.0"})
with urllib.request.urlopen(req, timeout=30) as resp:
    xml = resp.read().decode("utf-8-sig")

urls = re.findall(r"<loc>(https://sintech\.vn/collections/[^<]+)</loc>", xml)

os.makedirs(OUT_DIR, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Categories"

header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

ws["A1"] = "Tên Category"
ws["B1"] = "Link"
for cell in (ws["A1"], ws["B1"]):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

rows = []
for url in urls:
    slug = url.rstrip("/").split("/collections/")[-1]
    name = slug_to_name(slug)
    rows.append((name, url))

rows.sort(key=lambda x: x[0].lower())

for i, (name, url) in enumerate(rows, start=2):
    ws.cell(row=i, column=1, value=name)
    ws.cell(row=i, column=2, value=url)
    ws.cell(row=i, column=2).font = Font(color="0563C1", underline="single")

ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 70
ws.freeze_panes = "A2"

wb.save(OUT_FILE)
print(f"OK -> {OUT_FILE}")
print(f"Total rows: {len(rows)}")
