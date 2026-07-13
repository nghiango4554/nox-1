"""Tim trang TUNG CO TRAFFIC ma nay DA CHET (404) -> mat traffic + can 301 gap.

Boi canh (vo cho biet 13/7/2026): site truoc day manh nho cac bai huong dan cai dat
phan mem crack. Da xoa cac bai do NHUNG QUEN 301 -> URL chet, traffic sap.

Cach lam: GSC 16 thang lay MOI URL tung co hien thi -> curl kiem tra status hien tai
-> URL nao 404/410 ma tung co click/imp = thiet hai thuc + can 301 ve trang thay the.

Chay:  py -3.12 _scripts/dead_page_traffic.py
"""
import datetime
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TOKEN = r"C:\Users\NGHIANGO\.openclaw\workspace\gsc\token.json"
OUT = Path(r"C:\Users\NGHIANGO\.openclaw\workspace\nox-outputs")
UA = {"User-Agent": "Mozilla/5.0 (SintechHub audit)"}


def gsc_pages(svc, days, label):
    end = datetime.date.today() - datetime.timedelta(days=2)
    start = end - datetime.timedelta(days=days - 1)
    rows = svc.searchanalytics().query(siteUrl="sc-domain:sintech.vn", body={
        "startDate": start.isoformat(), "endDate": end.isoformat(),
        "dimensions": ["page"], "rowLimit": 25000,
    }).execute().get("rows", [])
    print(f"  {label}: {len(rows)} URL có hiển thị ({start} -> {end})")
    return {r["keys"][0]: {"clk": r["clicks"], "imp": r["impressions"],
                           "pos": round(r["position"], 1)} for r in rows}


def status(u):
    try:
        r = requests.get(u, headers=UA, timeout=25, allow_redirects=True)
        return u, r.status_code, r.url
    except Exception as e:
        return u, -1, f"{type(e).__name__}"


def main():
    creds = Credentials.from_authorized_user_file(TOKEN, ["https://www.googleapis.com/auth/webmasters"])
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    svc = build("searchconsole", "v1", credentials=creds, cache_discovery=False)

    print("Kéo dữ liệu GSC...")
    old = gsc_pages(svc, 480, "16 tháng")   # GSC giu toi da ~16 thang
    new = gsc_pages(svc, 28, "28 ngày gần đây")

    print(f"\nKiểm tra HTTP status {len(old)} URL...")
    with ThreadPoolExecutor(max_workers=10) as ex:
        res = list(ex.map(status, old))

    dead, alive_lost = [], []
    for u, code, final in res:
        d = old[u]
        if code in (404, 410, -1):
            dead.append({**d, "url": u, "code": code,
                         "imp_now": new.get(u, {}).get("imp", 0)})
        elif d["clk"] >= 20 and new.get(u, {}).get("clk", 0) == 0:
            alive_lost.append({**d, "url": u, "code": code})

    dead.sort(key=lambda x: -x["clk"])
    alive_lost.sort(key=lambda x: -x["clk"])

    tot_clk = sum(d["clk"] for d in dead)
    print(f"\n=== {len(dead)} URL CHẾT (404/410) mà từng có traffic ===")
    print(f"  Tổng click đã mất: {tot_clk} click/16 tháng\n")
    for d in dead[:30]:
        print(f"  {d['clk']:>5} click · {d['imp']:>6} imp · HTTP {d['code']} — {d['url'].replace('https://sintech.vn','')[:66]}")

    print(f"\n=== {len(alive_lost)} URL CÒN SỐNG nhưng MẤT SẠCH click (≥20 click cũ → 0 click nay) ===")
    for d in alive_lost[:20]:
        print(f"  {d['clk']:>5} click cũ · hạng cũ {d['pos']} — {d['url'].replace('https://sintech.vn','')[:66]}")

    wb = Workbook()
    ws = wb.active
    ws.title = "URL chết từng có traffic"
    ws.append(["Click (16 tháng)", "Hiển thị", "Hạng cũ", "HTTP", "URL", "Gợi ý 301 về"])
    for d in dead:
        ws.append([d["clk"], d["imp"], d["pos"], d["code"], d["url"], ""])
    ws2 = wb.create_sheet("URL sống nhưng mất click")
    ws2.append(["Click cũ", "Hiển thị", "Hạng cũ", "HTTP", "URL"])
    for d in alive_lost:
        ws2.append([d["clk"], d["imp"], d["pos"], d["code"], d["url"]])
    for w, widths in ((ws, [16, 11, 9, 8, 78, 60]), (ws2, [10, 11, 9, 8, 78])):
        for c in w[1]:
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for i, x in enumerate(widths, 1):
            w.column_dimensions[get_column_letter(i)].width = x
        w.freeze_panes = "A2"
    f = OUT / f"URL_chet_mat_traffic_{datetime.date.today():%Y%m%d}.xlsx"
    wb.save(f)
    print(f"\n[XONG] {f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
