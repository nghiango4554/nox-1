"""Ra soat TOAN BO redirect trong Haravan — tim cai TRO SAI HUONG.

4 kieu SAI (moi kieu deu lam mat thu hang):
1. DICH CHET   — 301 tro toi URL 404 => Google mat trang, khach vao trang loi. TE NHAT.
2. DICH LA TRANG DANH SACH / TRANG CHU — Google coi la soft-404, KHONG chuyen giao thu hang.
3. CHUOI REDIRECT — 301 -> 301 -> ... : loang tin hieu, cham. Nen tro THANG.
4. DICH LECH LOAI — /products/x -> /blogs/... hoac nguoc lai (thuong la tro bua).

Output: nox-outputs/Redirect_audit_<ngay>.xlsx + tab "Redirect audit" tren sheet Audit.

Chay:  py -3.12 _scripts/audit_redirects.py
"""
import datetime
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, r"C:\Users\NGHIANGO\.openclaw\workspace")
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import gsheet_client
import requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import haravan_blog as hb

SH = "13IDYcE2ZEUd64xK6dIN-P_3_4gIVyFcwdBjb8W8e-uU"
TAB = "Redirect audit"
OUT = Path(r"C:\Users\NGHIANGO\.openclaw\workspace\nox-outputs")
TOKEN = r"C:\Users\NGHIANGO\.openclaw\workspace\gsc\token.json"
UA = {"User-Agent": "Mozilla/5.0"}

LIST_PAGES = {"/blogs/huong-dan", "/blogs/news", "/blogs", "/collections", "/collections/all", "/", ""}


def all_redirects():
    H, B = hb._headers(), hb._base()
    out, page = [], 1
    while page <= 40:
        r = requests.get(f"{B}/redirects.json", headers=H, params={"limit": 250, "page": page}, timeout=40)
        if r.status_code != 200:
            time.sleep(5)
            continue
        d = r.json().get("redirects", [])
        if not d:
            break
        out += d
        page += 1
        time.sleep(0.5)
    seen, uniq = set(), []
    for x in out:
        if x["id"] not in seen:
            seen.add(x["id"])
            uniq.append(x)
    return uniq


def gsc_clicks():
    c = Credentials.from_authorized_user_file(TOKEN, ["https://www.googleapis.com/auth/webmasters"])
    if c.expired and c.refresh_token:
        c.refresh(Request())
    svc = build("searchconsole", "v1", credentials=c, cache_discovery=False)
    end = datetime.date.today() - datetime.timedelta(days=2)
    start = end - datetime.timedelta(days=479)
    rows = svc.searchanalytics().query(siteUrl="sc-domain:sintech.vn", body={
        "startDate": start.isoformat(), "endDate": end.isoformat(),
        "dimensions": ["page"], "rowLimit": 25000}).execute().get("rows", [])
    return {r["keys"][0].replace("https://sintech.vn", "").rstrip("/"): r["clicks"] for r in rows}


def head(u):
    """Trang thai cua DICH. Retry ky vi Haravan chan toc do (429)."""
    for i in range(6):
        try:
            r = requests.get("https://sintech.vn" + u, headers=UA, timeout=30, allow_redirects=False)
            if r.status_code == 429:
                time.sleep(8 + i * 4)
                continue
            return r.status_code, r.headers.get("Location", "").replace("https://sintech.vn", "")
        except Exception:
            time.sleep(3)
    return 429, ""


def main():
    reds = all_redirects()
    print(f"{len(reds)} redirect trong Haravan")
    clicks = gsc_clicks()

    targets = sorted({r["target"] for r in reds})
    print(f"{len(targets)} đích khác nhau — đang kiểm trạng thái từng đích...")
    with ThreadPoolExecutor(max_workers=3) as ex:
        st = dict(zip(targets, ex.map(head, targets)))

    rows, cnt = [], {}
    for r in reds:
        p, t = r["path"], r["target"]
        code, loc = st.get(t, (0, ""))
        clk = clicks.get(p.rstrip("/"), 0)

        if code in (404, 410):
            kind = "1. ĐÍCH CHẾT (404)"
        elif t.rstrip("/") in LIST_PAGES:
            kind = "2. Đích là trang danh sách/trang chủ"
        elif code in (301, 302):
            kind = "3. Chuỗi chuyển hướng (301 → 301)"
        elif p.startswith("/products/") and t.startswith("/blogs/"):
            kind = "4. Lệch loại (SP → bài viết)"
        elif p.startswith("/blogs/") and t.startswith("/products/"):
            kind = "4. Lệch loại (bài → SP)"
        elif code == 200:
            kind = "✔ OK"
        else:
            kind = f"? HTTP {code}"

        cnt[kind] = cnt.get(kind, 0) + 1
        if not kind.startswith("✔"):
            rows.append([kind, clk, p, t, code, loc])

    rows.sort(key=lambda x: (x[0], -x[1]))
    print("\n=== KẾT QUẢ ===")
    for k, v in sorted(cnt.items(), key=lambda kv: -kv[1]):
        print(f"  {v:>5} · {k}")

    print("\n=== Redirect SAI đáng chú ý (có click) ===")
    for k, clk, p, t, code, loc in [x for x in rows if x[1] > 0][:20]:
        print(f"  {clk:>5} click · [{k[:28]}] {p[:44]}")
        print(f"           → {t[:56]} (HTTP {code}{' → ' + loc[:24] if loc else ''})")

    wb = Workbook()
    ws = wb.active
    ws.title = "Redirect sai hướng"
    ws.append(["Kiểu sai", "Click (16 tháng)", "URL nguồn", "Đích hiện tại", "HTTP đích", "Đích nhảy tiếp"])
    for x in rows:
        ws.append(x)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, w in enumerate([34, 14, 58, 52, 10, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    f = OUT / f"Redirect_audit_{datetime.date.today():%Y%m%d}.xlsx"
    wb.save(f)

    svc = gsheet_client.get_service()
    meta = svc.spreadsheets().get(spreadsheetId=SH).execute()
    tabs = {s["properties"]["title"] for s in meta["sheets"]}
    if TAB not in tabs:
        svc.spreadsheets().batchUpdate(spreadsheetId=SH, body={"requests": [
            {"addSheet": {"properties": {"title": TAB, "gridProperties": {"frozenRowCount": 1}}}}]}).execute()
    else:
        svc.spreadsheets().values().clear(spreadsheetId=SH, range=f"'{TAB}'", body={}).execute()
    svc.spreadsheets().values().update(
        spreadsheetId=SH, range=f"'{TAB}'!A1", valueInputOption="RAW",
        body={"values": [["Kiểu sai", "Click", "URL nguồn", "Đích hiện tại", "HTTP đích", "Đích nhảy tiếp"]] + rows}
    ).execute()

    print(f"\n[XONG] {len(rows)} redirect có vấn đề")
    print(f"  Excel: {f}")
    print(f"  Sheet tab: {TAB}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
