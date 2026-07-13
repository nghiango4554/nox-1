"""Do ket qua FAQ v2 (nhom uu tien) -> tab moi tren Google Sheet + file Excel o Downloads.

Tab v1 cu ("FAQ Blog (duyệt)") GIU LAI lam tu lieu, nhung KHONG day len web:
Google da go FAQ rich result (7/5/2026) + v1 co cac loi da biet (ep 6 cau, khong bam key GSC,
loi viet "theo bai"). Ban dung de duyet la tab v2 nay.

Chay:  py -3.12 _scripts/faq_v2_to_sheet.py <file_v2.json>
"""
import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, r"C:\Users\NGHIANGO\.openclaw\workspace")
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import gsheet_client
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SHEET_ID = "13IDYcE2ZEUd64xK6dIN-P_3_4gIVyFcwdBjb8W8e-uU"
TAB = "FAQ v2 (ưu tiên)"
DL = Path(r"C:\Users\NGHIANGO\Downloads")
PALETTE_RGB = [(0.898, 0.945, 0.996), (0.925, 0.973, 0.925),
               (1.0, 0.953, 0.898), (0.960, 0.925, 0.976)]
PALETTE_HEX = ["E5F1FE", "ECF8EC", "FFF3E5", "F5ECF9"]
HEADER = ["STT", "Bài viết", "URL", "Lượt hiển thị", "Key chính (GSC)", "#",
          "Câu hỏi", "Câu trả lời", "Duyệt (OK/Bỏ)", "handle", "article_id", "blog_id"]


def rows_of(data):
    out = []
    for i, r in enumerate(data, 1):
        for k, f in enumerate(r["faqs"]):
            out.append([i if k == 0 else "", r["title"] if k == 0 else "", r["url"] if k == 0 else "",
                        r["imp"] if k == 0 else "", r.get("key_chinh", "") if k == 0 else "",
                        k + 1, f["q"], f["a"], "", r["handle"], str(r["id"]), str(r["blog_id"])])
    return out


def to_sheet(rows, n_art):
    svc = gsheet_client.get_service()
    meta = svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    tabs = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta["sheets"]}

    if TAB in tabs:  # xoa sach de ghi lai (chay lai khong nhan doi)
        sid = tabs[TAB]
        svc.spreadsheets().values().clear(spreadsheetId=SHEET_ID, range=f"'{TAB}'", body={}).execute()
    else:
        res = svc.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body={"requests": [
            {"addSheet": {"properties": {"title": TAB, "gridProperties": {"frozenRowCount": 1}}}}]}).execute()
        sid = res["replies"][0]["addSheet"]["properties"]["sheetId"]

    svc.spreadsheets().values().update(
        spreadsheetId=SHEET_ID, range=f"'{TAB}'!A1", valueInputOption="RAW",
        body={"values": [HEADER] + rows}).execute()

    reqs = [{"repeatCell": {
        "range": {"sheetId": sid, "startRowIndex": 0, "endRowIndex": 1},
        "cell": {"userEnteredFormat": {
            "backgroundColor": {"red": 0.12, "green": 0.31, "blue": 0.47},
            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            "horizontalAlignment": "CENTER", "wrapStrategy": "WRAP"}},
        "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,wrapStrategy)"}}]
    for i, w in enumerate([55, 300, 260, 90, 200, 40, 360, 520, 110, 220, 90, 90]):
        reqs.append({"updateDimensionProperties": {
            "range": {"sheetId": sid, "dimension": "COLUMNS", "startIndex": i, "endIndex": i + 1},
            "properties": {"pixelSize": w}, "fields": "pixelSize"}})

    blk, start, prev = -1, 0, None
    for idx, r in enumerate(rows):
        h = r[9]
        if h != prev:
            if prev is not None:
                reqs.append(_fill(sid, start, idx, blk))
            blk, start, prev = blk + 1, idx, h
    reqs.append(_fill(sid, start, len(rows), blk))

    for i in range(0, len(reqs), 100):
        svc.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body={"requests": reqs[i:i + 100]}).execute()
    print(f"[OK] Sheet tab '{TAB}': {n_art} bài · {len(rows)} câu")


def _fill(sid, s, e, blk):
    r, g, b = PALETTE_RGB[blk % 4]
    return {"repeatCell": {
        "range": {"sheetId": sid, "startRowIndex": s + 1, "endRowIndex": e + 1,
                  "startColumnIndex": 0, "endColumnIndex": len(HEADER)},
        "cell": {"userEnteredFormat": {
            "backgroundColor": {"red": r, "green": g, "blue": b},
            "textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 0}, "bold": False},
            "wrapStrategy": "WRAP", "verticalAlignment": "TOP"}},
        "fields": "userEnteredFormat(backgroundColor,textFormat,wrapStrategy,verticalAlignment)"}}


def to_excel(rows, n_art):
    wb = Workbook()
    ws = wb.active
    ws.title = "FAQ v2 (duyệt)"
    ws.append(HEADER)
    blk, prev = -1, None
    for r in rows:
        if r[9] != prev:
            blk, prev = blk + 1, r[9]
        ws.append(r)
        fill = PatternFill("solid", fgColor=PALETTE_HEX[blk % 4])
        for c in ws[ws.max_row]:
            c.fill, c.font = fill, Font(color="000000")
            c.alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([6, 42, 38, 11, 28, 5, 52, 70, 13, 30, 11, 11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"
    dv = DataValidation(type="list", formula1='"OK,Bỏ"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"I2:I{ws.max_row}")

    f = DL / f"FAQ_v2_uu_tien_{datetime.now():%Y%m%d}.xlsx"
    wb.save(f)
    print(f"[OK] Excel: {f}  ({n_art} bài · {len(rows)} câu)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("json_file")
    a = ap.parse_args()
    data = json.loads(Path(a.json_file).read_text(encoding="utf-8"))
    rows = rows_of(data)
    to_sheet(rows, len(data))
    to_excel(rows, len(data))
    return 0


if __name__ == "__main__":
    sys.exit(main())
