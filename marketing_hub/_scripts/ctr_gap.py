"""CTR-gap: tim truy van DA LEN TOP nhung KHONG AI BAM -> sua title/meta la an click ngay.

⚠️ BAY DA DINH (13/7/2026): KHONG duoc tinh CTR o cap (query x page).
Mot truy van co the hien NHIEU URL cua minh cung luc (sitelink) -> click don vao 1 URL,
cac URL con lai ghi 0 click => tao ra hang loat "CTR 0%" GIA.
Vi du: "sintech" tach theo trang thay 10 dong CTR ~0; nhung TONG thi 352 click/1048 imp = CTR 33.6%.
=> Phai do o cap QUERY (dimension=['query']), va loai tu khoa thuong hieu.

Logic: voi moi (query, page): so CTR thuc te vs CTR KY VONG theo hang.
CTR ky vong theo hang (benchmark nganh, uoc luong bao thu):
  1: 27% · 2: 15% · 3: 11% · 4: 8% · 5: 6% · 6: 5% · 7: 4% · 8: 3.5% · 9: 3% · 10: 2.5%
Click bi mat = imp * (ctr_ky_vong - ctr_thuc_te).

Xep theo CLICK BI MAT (khong phai theo CTR) -> uu tien viec dang tien nhat.

Chay:  py -3.12 _scripts/ctr_gap.py --days 90
"""
import argparse
import datetime
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TOKEN = r"C:\Users\NGHIANGO\.openclaw\workspace\gsc\token.json"
OUT = Path(r"C:\Users\NGHIANGO\.openclaw\workspace\nox-outputs")
EXPECTED = {1: .27, 2: .15, 3: .11, 4: .08, 5: .06, 6: .05, 7: .04, 8: .035, 9: .03, 10: .025}
BRAND = __import__("re").compile(r"sintech", __import__("re").I)  # brand: CTR 33.6%, khong phai gap


def expected_ctr(pos: float) -> float:
    return EXPECTED.get(int(round(pos)), 0.02 if pos <= 15 else 0.01)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=90)
    ap.add_argument("--min-imp", type=int, default=30, help="Bo qua truy van qua it hien thi")
    ap.add_argument("--max-pos", type=float, default=10.5, help="Chi xet truy van da o trang 1")
    a = ap.parse_args()

    creds = Credentials.from_authorized_user_file(TOKEN, ["https://www.googleapis.com/auth/webmasters"])
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    svc = build("searchconsole", "v1", credentials=creds, cache_discovery=False)
    end = datetime.date.today() - datetime.timedelta(days=2)
    start = end - datetime.timedelta(days=a.days - 1)
    print(f"GSC {start} -> {end} ({a.days} ngay)")

    # CTR THAT do o cap QUERY (xem canh bao dau file — tach theo page se ra CTR 0% GIA)
    rows = svc.searchanalytics().query(siteUrl="sc-domain:sintech.vn", body={
        "startDate": start.isoformat(), "endDate": end.isoformat(),
        "dimensions": ["query"], "rowLimit": 25000,
    }).execute().get("rows", [])

    # trang chu luc cua moi truy van (chi de biet SUA O DAU, khong dung tinh CTR)
    qp = {}
    for r in svc.searchanalytics().query(siteUrl="sc-domain:sintech.vn", body={
            "startDate": start.isoformat(), "endDate": end.isoformat(),
            "dimensions": ["query", "page"], "rowLimit": 25000}).execute().get("rows", []):
        k, u = r["keys"]
        cur = qp.get(k)
        if not cur or (r["clicks"], r["impressions"]) > (cur[1], cur[2]):
            qp[k] = (u, r["clicks"], r["impressions"])
    print(f"GSC tra {len(rows)} truy van\n")

    gaps = []
    for r in rows:
        q = r["keys"][0]
        if BRAND.search(q):   # tu khoa thuong hieu — CTR thuc 33.6%, khong sua bang title/meta
            continue
        p = qp.get(q, ("(khong ro)", 0, 0))[0]
        imp, clk, pos = r["impressions"], r["clicks"], r["position"]
        if imp < a.min_imp or pos > a.max_pos:
            continue
        ctr = clk / imp if imp else 0
        exp = expected_ctr(pos)
        lost = imp * (exp - ctr)
        if lost <= 0:
            continue
        gaps.append({"query": q, "page": p, "imp": imp, "clk": clk, "pos": round(pos, 1),
                     "ctr": round(ctr * 100, 2), "ctr_exp": round(exp * 100, 1),
                     "lost": round(lost, 1)})

    gaps.sort(key=lambda g: -g["lost"])
    tot_lost = sum(g["lost"] for g in gaps)
    zero = [g for g in gaps if g["clk"] == 0]

    print(f"=== {len(gaps)} truy van TOP 10 nhung CTR duoi ky vong ===")
    print(f"  Click bi mat uoc tinh: ~{int(tot_lost)} click/{a.days} ngay (~{int(tot_lost/a.days*30)}/thang)")
    print(f"  Trong do {len(zero)} truy van co hien thi nhung 0 CLICK\n")

    print(f"{'CLICK MẤT':>9} {'IMP':>6} {'HẠNG':>5} {'CTR':>6} {'KỲ VỌNG':>8}  TRUY VẤN")
    for g in gaps[:25]:
        print(f"{g['lost']:>9.0f} {g['imp']:>6} {g['pos']:>5} {g['ctr']:>5}% {g['ctr_exp']:>7}%  {g['query'][:44]}")

    # gom theo trang
    by_page = defaultdict(lambda: {"lost": 0, "imp": 0, "n": 0, "kw": []})
    for g in gaps:
        d = by_page[g["page"]]
        d["lost"] += g["lost"]
        d["imp"] += g["imp"]
        d["n"] += 1
        d["kw"].append(g["query"])
    top_pages = sorted(by_page.items(), key=lambda kv: -kv[1]["lost"])

    print("\n=== TOP 15 TRANG mat nhieu click nhat (sua title/meta la an ngay) ===")
    for u, d in top_pages[:15]:
        short = u.replace("https://sintech.vn", "")[:58]
        kw3 = ", ".join(d["kw"][:3])
        print(f"  ~{d['lost']:>5.0f} click mất · {d['n']:>2} kw · {d['imp']:>5} imp — {short}")
        print(f"        kw: {kw3}")

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "CTR gap theo truy vấn"
    ws.append(["Click bị mất", "Truy vấn", "Trang", "Hiển thị", "Click", "Hạng", "CTR %", "CTR kỳ vọng %"])
    for g in gaps:
        ws.append([g["lost"], g["query"], g["page"], g["imp"], g["clk"], g["pos"], g["ctr"], g["ctr_exp"]])

    ws2 = wb.create_sheet("CTR gap theo trang")
    ws2.append(["Click bị mất", "Trang", "Số truy vấn", "Hiển thị", "Truy vấn tiêu biểu"])
    for u, d in top_pages:
        ws2.append([round(d["lost"], 1), u, d["n"], d["imp"], " · ".join(d["kw"][:6])])

    for w, widths in ((ws, [12, 46, 60, 10, 8, 7, 8, 12]), (ws2, [12, 60, 11, 10, 80])):
        for c in w[1]:
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for i, x in enumerate(widths, 1):
            w.column_dimensions[get_column_letter(i)].width = x
        w.freeze_panes = "A2"
        w.auto_filter.ref = f"A1:{get_column_letter(w.max_column)}{w.max_row}"

    f = OUT / f"CTR_gap_{datetime.date.today():%Y%m%d}.xlsx"
    wb.save(f)
    print(f"\n[XONG] {f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
