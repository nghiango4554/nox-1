# -*- coding: utf-8 -*-
"""Gen title/meta phong cách mới (hook mạnh) cho các bài trong sheet 'Cần sửa title-meta'.
Nguồn = danh sách flagged (title ngoài 45-60 / meta ngoài 140-165), thứ tự y sheet Cần sửa.
Chỉ gen ĐÚNG field đang lỗi (giữ field ổn). START/COUNT để chạy theo lô."""
import sys, io, re, sqlite3, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import collection_content_writer as ccw
import ai_provider
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAIN = r'C:\Users\NGHIANGO\.openclaw\workspace\nox-outputs\collections-overview.xlsx'
SHEET = "Gen phong cách mới"
START, COUNT = 0, 5          # 5 bài đầu sheet Cần sửa
T_LO, T_HI = 45, 60
M_LO, M_HI = 140, 165
SUF = re.compile(r'\s*[–\-|]\s*Sintech(\.vn)?\s*$', re.I)

STYLE = """

PHONG CÁCH LẦN NÀY — HOOK MẠNH, GIỌNG BÁN HÀNG (KHÔNG mô tả hiền lành):
1. DẪN bằng HOOK MẠNH ngay đầu meta: chạm NỖI ĐAU/nhu cầu khách rồi phang giải pháp dứt khoát.
   VD "Máy nóng, ồn, chậm? Vệ sinh PC tận nơi, sạch bụi - mát tức thì...". CẤM mở nhạt kiểu "Dịch vụ/Sản phẩm X tại...".
2. Tín hiệu thương mại CÓ THẬT của Sintech, phang MẠNH, hợp loại danh mục:
   - Sản phẩm giá trị cao (PC/laptop/màn hình/nguồn/linh kiện chính): "TRẢ GÓP 0%", "GIÁ TỐT", "CHÍNH HÃNG", "BẢO HÀNH CHÍNH HÃNG", "GIAO NHANH".
   - Phụ kiện/gear nhỏ: "CHÍNH HÃNG", "GIÁ TỐT", "GIAO NHANH", "BẢO HÀNH".
   - DỊCH VỤ (vệ sinh/sửa chữa/cài đặt/nâng cấp): "TẬN NƠI", "THỢ ĐẾN TẬN NHÀ", "NHANH - GỌN", "UY TÍN", "BẢO HÀNH DỊCH VỤ", "Q7 / TP.HCM". TUYỆT ĐỐI KHÔNG "trả góp" cho dịch vụ.
3. Câu CỤM NGẮN, DỨT KHOÁT. BỎ từ lề mề: "hỗ trợ", "có thể", "kiểm tra rõ lỗi", "tư vấn đúng nhu cầu".
4. CTA MẠNH cuối meta: DỊCH VỤ → "GỌI NGAY 0911 713 000" / "ĐẶT LỊCH NGAY"; SẢN PHẨM → "MUA NGAY" / "XEM GIÁ NGAY".
5. TITLE phải có lực (lợi ích + 1 hook), không liệt kê suông; viết câu thường (KHÔNG Viết Hoa Từng Từ).
6. GIỮ ĐỘ DÀI: TITLE 50-58, META 150-160. KHÔNG bịa số liệu/cam kết không có thật — MẠNH Ở GIỌNG."""

SYS = ccw._TITLE_META_SYSTEM_PROMPT + STYLE

c = sqlite3.connect('data/posts.db')
rows = c.execute("""
    SELECT sp.url, COALESCE(sp.title,''), COALESCE(sp.meta_desc,''),
           cj.haravan_id, cj.collection_title, COALESCE(sp.h1,''), cj.tier_level, cj.tier1_name
    FROM seo_pages sp
    LEFT JOIN collection_jobs cj
      ON cj.handle = replace(replace(sp.url,'https://sintech.vn/collections/',''),'/','')
    WHERE sp.url LIKE '%/collections/%' AND sp.status_code=200
      AND COALESCE(sp.indexable,1)=1 AND COALESCE(sp.excluded_from_audit,0)=0
""").fetchall()

flagged = []
for url, title, meta, hid, cj_title, h1, lvl, t1 in rows:
    bare = SUF.sub('', title).strip()
    tl, ml = len(bare), len(meta)
    t_bad = tl < T_LO or tl > T_HI
    m_bad = ml < M_LO or ml > M_HI
    if not (t_bad or m_bad):
        continue
    field = 'both' if (t_bad and m_bad) else ('title' if t_bad else 'meta')
    flagged.append({'url': url, 'bare': bare, 'meta': meta, 'tl': tl, 'ml': ml,
                    'field': field, 't_bad': t_bad, 'm_bad': m_bad, 'hid': hid,
                    'lvl': int(lvl) if lvl else 9, 't1': t1 or 'zzz',
                    'name': (cj_title or '').strip() or h1 or url.rstrip('/').split('/collections/')[-1]})
# Sort y sheet Cần sửa: lỗi cả 2 trước, rồi t1/lvl/tên
flagged.sort(key=lambda r: (not (r['t_bad'] and r['m_bad']), r['t1'], r['lvl'], r['name'].lower()))
pick = flagged[START:START + COUNT]

def gen(r):
    sp_names = []
    if r['hid']:
        try:
            hv = ccw.fetch_real_products(r['hid'])
            if hv.get('ok'): sp_names = hv.get('names', [])
        except Exception: pass
    field = r['field']
    if field == 'title':   focus = 'CHỈ gen lại TITLE'; schema = '{"title":"..."}'
    elif field == 'meta':  focus = 'CHỈ gen lại META'; schema = '{"meta":"..."}'
    else:                  focus = 'Gen lại CẢ title VÀ meta'; schema = '{"title":"...","meta":"..."}'
    usr = f"""COLLECTION cần gen lại ({field}):
- Tên: {r['name']}
- URL: {r['url']}
- Top SP ({len(sp_names)} mẫu): {', '.join(sp_names[:6]) if sp_names else '(dịch vụ / không có SP)'}
- TITLE cũ: {r['bare']}
- META cũ:  {r['meta']}
→ {focus}, áp PHONG CÁCH LẦN NÀY, đổi góc nhìn. Trả JSON {schema} duy nhất."""
    try:
        raw = ai_provider.call_ai(SYS, usr, timeout=90)
    except Exception as e:
        return '', '', str(e)
    t = re.sub(r'^```(?:json)?\s*', '', raw.strip()); t = re.sub(r'\s*```\s*$', '', t)
    try:
        d = json.loads(t)
    except Exception:
        m = re.search(r'\{[\s\S]*\}', t); d = json.loads(m.group(0)) if m else {}
    return (d.get('title', '') or '').strip(), (d.get('meta', '') or '').strip(), ''

results = []
for i, r in enumerate(pick):
    nt, nm, err = gen(r)
    results.append({**r, 'nt': nt, 'nm': nm, 'err': err, 'stt': START + i + 1})
    print(f"  [{START+i+1}] {r['name'][:26]:26} {r['field']:5} -> T[{len(nt) if nt else '-'}] M[{len(nm) if nm else '-'}] {err}")

# ---- ghi sheet ----
wb = openpyxl.load_workbook(MAIN)
HF = PatternFill("solid", fgColor="2E7D32"); WF = Font(bold=True, color="FFFFFF")
OLDF = PatternFill("solid", fgColor="FCE4D6"); NEWF = PatternFill("solid", fgColor="E2EFDA")
GOOD = PatternFill("solid", fgColor="C6EFCE"); WARN = PatternFill("solid", fgColor="FFEB9C")
thin = Side(style='thin', color="D9D9D9"); B = Border(thin, thin, thin, thin)
WRAP = Alignment(wrap_text=True, vertical='top'); CEN = Alignment('center', 'center', wrap_text=True)
HEAD = ['STT', 'Tên collection', 'Field', 'Title CŨ', 'CŨ\nlen', 'Title MỚI', 'MỚI\nlen',
        'Meta CŨ', 'CŨ\nlen', 'Meta MỚI', 'MỚI\nlen', 'Link']
if START == 0 and SHEET in wb.sheetnames:
    del wb[SHEET]
if SHEET in wb.sheetnames:
    ws = wb[SHEET]; start_row = ws.max_row + 1
else:
    ws = wb.create_sheet(SHEET, index=1)
    for ci, h in enumerate(HEAD, 1):
        cell = ws.cell(1, ci, h); cell.fill = HF; cell.font = WF; cell.alignment = CEN; cell.border = B
    for ci, w in enumerate([5, 22, 7, 40, 6, 42, 6, 50, 6, 50, 6, 26], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30; ws.freeze_panes = 'B2'
    start_row = 2
for i, r in enumerate(results):
    rr = start_row + i; nt, nm = r['nt'], r['nm']
    show_t = nt or ('(giữ nguyên)' if r['field'] == 'meta' else '')
    show_m = nm or ('(giữ nguyên)' if r['field'] == 'title' else '')
    vals = [r['stt'], r['name'], r['field'], r['bare'], r['tl'], show_t, len(nt) if nt else '',
            r['meta'], r['ml'], show_m, len(nm) if nm else '', r['url']]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(rr, ci, v); cell.border = B
        cell.alignment = CEN if ci in (1, 3, 5, 7, 9, 11) else WRAP
        if ci in (4, 8): cell.fill = OLDF
        if ci in (6, 10): cell.fill = NEWF
    if nt: ws.cell(rr, 7).fill = GOOD if 48 <= len(nt) <= 58 else WARN
    if nm: ws.cell(rr, 11).fill = GOOD if 140 <= len(nm) <= 160 else WARN
    ws.cell(rr, 12).hyperlink = r['url']; ws.cell(rr, 12).font = Font(color="0563C1", underline="single")
    ws.row_dimensions[rr].height = 80
wb.save(MAIN)
print(f"\nĐã ghi {len(results)} bài (STT {START+1}-{START+len(results)} của sheet Cần sửa) vào '{SHEET}' | tổng {ws.max_row-1} dòng")
