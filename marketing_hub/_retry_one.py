# -*- coding: utf-8 -*-
"""Retry gen 1 bài (theo STT trong sheet Cần sửa) + đẩy thẳng lên Google Sheet Link-Cate."""
import sys, io, os, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import collection_content_writer as ccw
import ai_provider
import openpyxl
sys.path.insert(0, r'C:\Users\NGHIANGO\.openclaw\workspace')
from gsheet_client import get_service

STT = 53
MAIN = r'C:\Users\NGHIANGO\.openclaw\workspace\nox-outputs\collections-overview.xlsx'
GSID = '1B0WtpBeeST0Pyw5Z9R08r00A_MUMbdi9YDqnJGKkdyM'
T_LO, T_HI, M_LO, M_HI = 45, 60, 140, 165
SUF = re.compile(r'\s*[–\-|]\s*Sintech(\.vn)?\s*$', re.I)
norm = lambda u: (u or '').strip().rstrip('/').lower()

STYLE = """

PHONG CÁCH — HOOK MẠNH, GIỌNG BÁN HÀNG:
1. CHỈ META dẫn bằng HOOK NỖI ĐAU/nhu cầu rồi phang giải pháp. CẤM mở meta nhạt.
2. Tín hiệu thật hợp loại: SP giá trị cao → TRẢ GÓP 0%/GIÁ TỐT/CHÍNH HÃNG/BẢO HÀNH CHÍNH HÃNG/GIAO NHANH; phụ kiện → CHÍNH HÃNG/GIÁ TỐT/GIAO NHANH/BẢO HÀNH.
3. Cụm ngắn, dứt khoát, bỏ lề mề.
4. CTA mạnh cuối meta: SP → MUA NGAY/XEM GIÁ NGAY.
5. TITLE: TUYỆT ĐỐI KHÔNG câu hỏi/"?"; khẳng định lợi ích + hook, câu thường.
6. TITLE 50-58, META 150-160. Không bịa số liệu."""
SYS = ccw._TITLE_META_SYSTEM_PROMPT + STYLE

src = openpyxl.load_workbook(MAIN, data_only=True)['Cần sửa title-meta']
row = STT + 1
name = src.cell(row, 3).value; bare = src.cell(row, 4).value or ''
tl = src.cell(row, 5).value or 0; meta = src.cell(row, 6).value or ''; ml = src.cell(row, 7).value or 0
url = src.cell(row, 9).value
t_bad = tl < T_LO or tl > T_HI; m_bad = ml < M_LO or ml > M_HI
field = 'both' if (t_bad and m_bad) else ('title' if t_bad else 'meta')
print(f"STT{STT}: {name} | field={field} | {url}")

focus = {'title': 'CHỈ gen lại TITLE', 'meta': 'CHỈ gen lại META', 'both': 'Gen lại CẢ title VÀ meta'}[field]
schema = {'title': '{"title":"..."}', 'meta': '{"meta":"..."}', 'both': '{"title":"...","meta":"..."}'}[field]
usr = f"""COLLECTION cần gen lại ({field}):
- Tên: {name}
- URL: {url}
- TITLE cũ: {bare}
- META cũ:  {meta}
→ {focus}, áp PHONG CÁCH. Trả JSON {schema} duy nhất."""
raw = ai_provider.call_ai(SYS, usr, timeout=120)
t = re.sub(r'^```(?:json)?\s*', '', raw.strip()); t = re.sub(r'\s*```\s*$', '', t)
try: d = json.loads(t)
except Exception:
    m = re.search(r'\{[\s\S]*\}', t); d = json.loads(m.group(0))
nt = (d.get('title', '') or '').strip(); nm = (d.get('meta', '') or '').strip()
if nm and not (M_LO <= len(nm) <= 160):
    nm, _ = ccw._fix_meta_length(nm, name, max_ai_retries=1)
print(f"  -> TITLE[{len(nt)}]: {nt}")
print(f"  -> META [{len(nm)}]: {nm}")

# push thẳng lên Link-Cate
svc = get_service()
lc = svc.spreadsheets().values().get(spreadsheetId=GSID, range="'Link-Cate'!A1:B1000").execute().get('values', [])
rownum = next((i for i, r in enumerate(lc, 1) if len(r) > 1 and norm(r[1]) == norm(url)), None)
if rownum:
    svc.spreadsheets().values().update(spreadsheetId=GSID, range=f"'Link-Cate'!C{rownum}:F{rownum}",
        valueInputOption='RAW', body={'values': [[nt, len(nt) if nt else '', nm, len(nm) if nm else '']]}).execute()
    print(f"  ✓ đẩy lên Link-Cate dòng {rownum}")
else:
    print("  ✗ không tìm thấy link trong Link-Cate")
