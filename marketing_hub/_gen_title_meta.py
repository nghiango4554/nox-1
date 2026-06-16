# -*- coding: utf-8 -*-
"""Gen lại title/meta collection lệch chuẩn -> thêm sheet vào collections-overview.xlsx.
LIMIT = 5  -> gen thử 5 bài đại diện.
LIMIT = None -> gen TẤT CẢ collection lệch chuẩn.
"""
import sys, io, re, sqlite3
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import collection_content_writer as ccw
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LIMIT = 5                          # đổi None để chạy hết 75
SHEET = "Gen thử 5" if LIMIT else "Gen lại title-meta"
MAIN = r'C:\Users\NGHIANGO\.openclaw\workspace\nox-outputs\collections-overview.xlsx'
T_LO, T_HI = 45, 60
M_LO, M_HI = 140, 165
SUF = re.compile(r'\s*[–\-|]\s*Sintech(\.vn)?\s*$', re.I)

c = sqlite3.connect('data/posts.db')
rows = c.execute("""
    SELECT sp.url, COALESCE(sp.title,''), COALESCE(sp.meta_desc,''),
           cj.haravan_id, cj.collection_title, COALESCE(sp.h1,''), cj.tier_level, cj.tier1_name
    FROM seo_pages sp
    LEFT JOIN collection_jobs cj
      ON cj.handle = replace(replace(sp.url,'https://sintech.vn/collections/',''),'/','')
    WHERE sp.url LIKE '%/collections/%' AND sp.status_code=200
      AND COALESCE(sp.indexable,1)=1 AND COALESCE(sp.excluded_from_audit,0)=0
""").fetchall()

flagged = []
for url, title, meta, hid, cj_title, h1, lvl, t1 in rows:
    bare = SUF.sub('', title).strip()
    tl, ml = len(bare), len(meta)
    t_bad = tl < T_LO or tl > T_HI
    m_bad = ml < M_LO or ml > M_HI
    if not (t_bad or m_bad):
        continue
    handle = url.rstrip('/').split('/collections/')[-1]
    field = 'both' if (t_bad and m_bad) else ('title' if t_bad else 'meta')
    flagged.append({'handle': handle, 'url': url, 'bare': bare, 'meta': meta, 'tl': tl, 'ml': ml,
                    'field': field, 'hid': hid, 'tier': (f"T{lvl}·{t1}" if lvl else "(chưa tầng)"),
                    'name': (cj_title or '').strip() or h1 or handle})

if LIMIT:  # 5 đại diện: 1 both, 2 title, 2 meta
    pick = []
    def take(pred, n):
        for r in sorted(flagged, key=lambda x: -(abs(x['tl']-53)+abs(x['ml']-152))):
            if r in pick: continue
            if pred(r): pick.append(r); n -= 1
            if n == 0: break
    take(lambda r: r['field']=='both', 1)
    take(lambda r: r['field']=='title', 2)
    take(lambda r: r['field']=='meta', 2)
else:
    pick = sorted(flagged, key=lambda r: (r['field'] != 'both', r['name'].lower()))

results = []
for idx, r in enumerate(pick, 1):
    sp_names = []
    if r['hid']:
        try:
            hv = ccw.fetch_real_products(r['hid'])
            if hv.get('ok'): sp_names = hv.get('names', [])
        except Exception: pass
    res = ccw.gen_title_meta_only(r['url'], r['name'], sp_names=sp_names,
                                  existing_title=r['bare'], existing_meta=r['meta'], field=r['field'])
    nt = res.get('title', '') if res.get('ok') else ''
    nm = res.get('meta', '') if res.get('ok') else ''
    results.append({**r, 'new_title': nt, 'new_meta': nm, 'err': '' if res.get('ok') else res.get('error', '')})
    print(f"  [{idx}/{len(pick)}] {r['handle']:26} {r['field']:5} -> T[{len(nt) if nt else '-'}] M[{len(nm) if nm else '-'}] {res.get('error','')}")

# ---- ghi sheet vào workbook chung ----
wb = openpyxl.load_workbook(MAIN)
if SHEET in wb.sheetnames: del wb[SHEET]
ws = wb.create_sheet(SHEET)
HF = PatternFill("solid", fgColor="1F4E79"); WF = Font(bold=True, color="FFFFFF")
OLDF = PatternFill("solid", fgColor="FCE4D6"); NEWF = PatternFill("solid", fgColor="E2EFDA")
GOOD = PatternFill("solid", fgColor="C6EFCE"); WARN = PatternFill("solid", fgColor="FFEB9C")
thin = Side(style='thin', color="D9D9D9"); B = Border(thin, thin, thin, thin)
WRAP = Alignment(wrap_text=True, vertical='top'); CEN = Alignment('center', 'center', wrap_text=True)
HEAD = ['STT', 'Tên collection', 'Field', 'Title CŨ', 'CŨ\nlen', 'Title MỚI', 'MỚI\nlen',
        'Meta CŨ', 'CŨ\nlen', 'Meta MỚI', 'MỚI\nlen', 'Link']
for ci, h in enumerate(HEAD, 1):
    cell = ws.cell(1, ci, h); cell.fill = HF; cell.font = WF; cell.alignment = CEN; cell.border = B
for i, r in enumerate(results, 1):
    rr = i + 1; nt, nm = r['new_title'], r['new_meta']
    vals = [i, r['name'], r['field'], r['bare'], r['tl'], nt or '(giữ nguyên)', len(nt) if nt else '',
            r['meta'], r['ml'], nm or '(giữ nguyên)', len(nm) if nm else '', r['url']]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(rr, ci, v); cell.border = B
        cell.alignment = CEN if ci in (1, 3, 5, 7, 9, 11) else WRAP
        if ci in (4, 8): cell.fill = OLDF
        if ci in (6, 10): cell.fill = NEWF
    if nt: ws.cell(rr, 7).fill = GOOD if 48 <= len(nt) <= 58 else WARN
    if nm: ws.cell(rr, 11).fill = GOOD if 140 <= len(nm) <= 160 else WARN
    ws.cell(rr, 12).hyperlink = r['url']; ws.cell(rr, 12).font = Font(color="0563C1", underline="single")
for ci, w in enumerate([5, 22, 7, 40, 6, 40, 6, 50, 6, 50, 6, 28], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 30
for rr in range(2, len(results) + 2): ws.row_dimensions[rr].height = 70
ws.freeze_panes = 'B2'
wb.save(MAIN)
print(f"\nĐã thêm sheet '{SHEET}' vào {MAIN} | sheets: {wb.sheetnames}")
