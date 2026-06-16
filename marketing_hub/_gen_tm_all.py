# -*- coding: utf-8 -*-
"""Gen title/meta phong cách mới cho TẤT CẢ bài trong 'Cần sửa title-meta'.
3 luồng + ép meta 150-160 + lưu workbook mỗi 10 bài (resilient)."""
import sys, io, re, sqlite3, json, threading
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import collection_content_writer as ccw
import ai_provider
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

MAIN = r'C:\Users\NGHIANGO\.openclaw\workspace\nox-outputs\collections-overview.xlsx'
SHEET = "Gen phong cách mới"
T_LO, T_HI = 45, 60
M_LO, M_HI = 140, 165
SUF = re.compile(r'\s*[–\-|]\s*Sintech(\.vn)?\s*$', re.I)

STYLE = """

PHONG CÁCH LẦN NÀY — HOOK MẠNH, GIỌNG BÁN HÀNG (KHÔNG mô tả hiền lành):
1. CHỈ META mới dẫn bằng HOOK NỖI ĐAU/nhu cầu khách rồi phang giải pháp. CẤM mở meta nhạt "Dịch vụ/Sản phẩm X tại...".
2. Tín hiệu thương mại CÓ THẬT của Sintech, hợp loại danh mục:
   - SP giá trị cao (PC/laptop/màn hình/nguồn/VGA/CPU/mainboard): "TRẢ GÓP 0%", "GIÁ TỐT", "CHÍNH HÃNG", "BẢO HÀNH CHÍNH HÃNG", "GIAO NHANH".
   - Phụ kiện/gear nhỏ: "CHÍNH HÃNG", "GIÁ TỐT", "GIAO NHANH", "BẢO HÀNH".
   - DỊCH VỤ (vệ sinh/sửa chữa/cài đặt/nâng cấp): "TẬN NƠI", "THỢ ĐẾN TẬN NHÀ", "NHANH - GỌN", "UY TÍN", "BẢO HÀNH DỊCH VỤ", "Q7/TP.HCM". KHÔNG "trả góp" cho dịch vụ.
3. Câu CỤM NGẮN, DỨT KHOÁT. BỎ lề mề: "hỗ trợ", "có thể", "kiểm tra rõ lỗi".
4. CTA MẠNH cuối meta: DỊCH VỤ → "GỌI NGAY 0911 713 000"/"ĐẶT LỊCH NGAY"; SP → "MUA NGAY"/"XEM GIÁ NGAY".
5. TITLE: TUYỆT ĐỐI KHÔNG dùng câu hỏi/dấu "?". Viết KHẲNG ĐỊNH lợi ích + 1 hook thương mại, dứt khoát, câu thường (KHÔNG Viết Hoa Từng Từ). VD "Màn hình Asus chính hãng, lên hình mượt, giá tốt".
6. ĐỘ DÀI: TITLE 50-58, META 150-160. KHÔNG bịa số liệu/cam kết."""
SYS = ccw._TITLE_META_SYSTEM_PROMPT + STYLE

# Nguồn = sheet 'Cần sửa title-meta' (snapshot, KHÔNG phụ thuộc seo_pages đang re-crawl)
c = sqlite3.connect('data/posts.db')
def get_hid(handle):
    row = c.execute("select haravan_id from collection_jobs where handle=?", (handle,)).fetchone()
    return row[0] if row else None

src = openpyxl.load_workbook(MAIN, data_only=True)['Cần sửa title-meta']
flagged = []
for r in range(2, src.max_row + 1):
    name = src.cell(r, 3).value
    bare = src.cell(r, 4).value or ''
    tl = src.cell(r, 5).value or 0
    meta = src.cell(r, 6).value or ''
    ml = src.cell(r, 7).value or 0
    url = src.cell(r, 9).value
    if not url: continue
    t_bad = tl < T_LO or tl > T_HI; m_bad = ml < M_LO or ml > M_HI
    field = 'both' if (t_bad and m_bad) else ('title' if t_bad else 'meta')
    handle = url.rstrip('/').split('/collections/')[-1]
    flagged.append({'url': url, 'bare': bare, 'meta': meta, 'tl': tl, 'ml': ml, 'field': field,
                    't_bad': t_bad, 'm_bad': m_bad, 'hid': get_hid(handle), 'name': name,
                    'stt': r - 1})
print(f"Gen ALL: {len(flagged)} bài (nguồn: sheet Cần sửa), 3 luồng...")

def gen(r):
    sp_names = []
    if r['hid']:
        try:
            hv = ccw.fetch_real_products(r['hid'])
            if hv.get('ok'): sp_names = hv.get('names', [])
        except Exception: pass
    f = r['field']
    focus = {'title': 'CHỈ gen lại TITLE', 'meta': 'CHỈ gen lại META', 'both': 'Gen lại CẢ title VÀ meta'}[f]
    schema = {'title': '{"title":"..."}', 'meta': '{"meta":"..."}', 'both': '{"title":"...","meta":"..."}'}[f]
    usr = f"""COLLECTION cần gen lại ({f}):
- Tên: {r['name']}
- URL: {r['url']}
- Top SP ({len(sp_names)}): {', '.join(sp_names[:6]) if sp_names else '(dịch vụ/không SP)'}
- TITLE cũ: {r['bare']}
- META cũ:  {r['meta']}
→ {focus}, áp PHONG CÁCH LẦN NÀY. Trả JSON {schema} duy nhất."""
    try:
        raw = ai_provider.call_ai(SYS, usr, timeout=90)
    except Exception as e:
        return r['stt'], {**r, 'nt': '', 'nm': '', 'err': str(e)[:60]}
    t = re.sub(r'^```(?:json)?\s*', '', raw.strip()); t = re.sub(r'\s*```\s*$', '', t)
    try: d = json.loads(t)
    except Exception:
        m = re.search(r'\{[\s\S]*\}', t); d = json.loads(m.group(0)) if m else {}
    nt = (d.get('title', '') or '').strip(); nm = (d.get('meta', '') or '').strip()
    if nm and not (M_LO <= len(nm) <= 160):   # ép meta về 140-160
        try: nm, _ = ccw._fix_meta_length(nm, r['name'], max_ai_retries=1)
        except Exception: pass
    return r['stt'], {**r, 'nt': nt, 'nm': nm, 'err': ''}

# ---- styles ----
HF = PatternFill("solid", fgColor="2E7D32"); WF = Font(bold=True, color="FFFFFF")
OLDF = PatternFill("solid", fgColor="FCE4D6"); NEWF = PatternFill("solid", fgColor="E2EFDA")
GOOD = PatternFill("solid", fgColor="C6EFCE"); WARN = PatternFill("solid", fgColor="FFEB9C")
thin = Side(style='thin', color="D9D9D9"); B = Border(thin, thin, thin, thin)
WRAP = Alignment(wrap_text=True, vertical='top'); CEN = Alignment('center', 'center', wrap_text=True)
HEAD = ['STT', 'Tên collection', 'Field', 'Title CŨ', 'CŨ\nlen', 'Title MỚI', 'MỚI\nlen',
        'Meta CŨ', 'CŨ\nlen', 'Meta MỚI', 'MỚI\nlen', 'Link']

def save(done_map):
    wb = openpyxl.load_workbook(MAIN)
    if SHEET in wb.sheetnames: del wb[SHEET]
    ws = wb.create_sheet(SHEET, index=1)
    for ci, h in enumerate(HEAD, 1):
        cell = ws.cell(1, ci, h); cell.fill = HF; cell.font = WF; cell.alignment = CEN; cell.border = B
    for ci, w in enumerate([5, 22, 7, 40, 6, 42, 6, 50, 6, 50, 6, 26], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30; ws.freeze_panes = 'B2'
    for rr, stt in enumerate(sorted(done_map), start=2):
        r = done_map[stt]; nt, nm = r['nt'], r['nm']
        show_t = nt or ('(giữ nguyên)' if r['field'] == 'meta' else ('LỖI: ' + r['err'] if r['err'] else ''))
        show_m = nm or ('(giữ nguyên)' if r['field'] == 'title' else '')
        vals = [r['stt'], r['name'], r['field'], r['bare'], r['tl'], show_t, len(nt) if nt else '',
                r['meta'], r['ml'], show_m, len(nm) if nm else '', r['url']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(rr, ci, v); cell.border = B
            cell.alignment = CEN if ci in (1, 3, 5, 7, 9, 11) else WRAP
            if ci in (4, 8): cell.fill = OLDF
            if ci in (6, 10): cell.fill = NEWF
        if nt: ws.cell(rr, 7).fill = GOOD if 48 <= len(nt) <= 58 else WARN
        if nm: ws.cell(rr, 11).fill = GOOD if 140 <= len(nm) <= 160 else WARN
        ws.cell(rr, 12).hyperlink = r['url']; ws.cell(rr, 12).font = Font(color="0563C1", underline="single")
        ws.row_dimensions[rr].height = 80
    for attempt in range(3):
        try: wb.save(MAIN); return True
        except PermissionError: pass
    # fallback nếu file bị khóa
    json.dump({str(k): {kk: vv for kk, vv in v.items() if kk in ('stt','name','nt','nm','err')}
               for k, v in done_map.items()}, open(MAIN + '.backup.json', 'w', encoding='utf-8'), ensure_ascii=False)
    return False

done = {}
with ThreadPoolExecutor(max_workers=3) as ex:
    futs = [ex.submit(gen, r) for r in flagged]
    n = 0
    for fut in as_completed(futs):
        stt, res = fut.result(); done[stt] = res; n += 1
        print(f"  [{n}/{len(flagged)}] STT{stt} {res['name'][:24]:24} T[{len(res['nt']) if res['nt'] else '-'}] M[{len(res['nm']) if res['nm'] else '-'}] {res['err']}")
        if n % 10 == 0: save(done)
ok = save(done)
errs = sum(1 for v in done.values() if v['err'])
print(f"\nXONG {len(done)} bài | lỗi: {errs} | saved={ok}")
