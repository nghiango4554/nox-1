# -*- coding: utf-8 -*-
"""MASTER: 1 file collections-overview.xlsx nhiều sheet con.
- Sheet "Tổng quan (211)": mọi thông số collection live.
- Sheet "Cần sửa title-meta": lọc title ngoài [45,60] / meta ngoài [140,165].
- Sheet "Gen title-meta" do script gen tự thêm vào CÙNG file này (append).
"""
import sys, io, re, sqlite3
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from seo import USER_AGENT, TIMEOUT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r'C:\Users\NGHIANGO\.openclaw\workspace\nox-outputs\collections-overview.xlsx'
T_LO, T_HI = 45, 60
M_LO, M_HI = 140, 165
WORD_FLOOR = 1000
SUF = re.compile(r'\s*[–\-|]\s*Sintech(\.vn)?\s*$', re.I)

# ---- styles ----
HF = PatternFill("solid", fgColor="1F4E79"); WF = Font(bold=True, color="FFFFFF")
BAD = PatternFill("solid", fgColor="FFC7CE"); WARN = PatternFill("solid", fgColor="FFEB9C")
ALT = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style='thin', color="D9D9D9"); B = Border(thin, thin, thin, thin)
WRAP = Alignment(wrap_text=True, vertical='top'); CEN = Alignment('center', 'center', wrap_text=True)

def style_header(ws, head, fill=HF):
    for ci, h in enumerate(head, 1):
        c = ws.cell(1, ci, h); c.fill = fill; c.font = WF; c.alignment = CEN; c.border = B

# ---- data ----
c = sqlite3.connect('data/posts.db')
rows = c.execute("""
    SELECT sp.url, COALESCE(sp.title,''), COALESCE(sp.meta_desc,''), COALESCE(sp.h1,''),
           cj.tier_level, cj.tier1_name, cj.collection_title
    FROM seo_pages sp
    LEFT JOIN collection_jobs cj
      ON cj.handle = replace(replace(sp.url,'https://sintech.vn/collections/',''),'/','')
    WHERE sp.url LIKE '%/collections/%' AND sp.status_code=200
      AND COALESCE(sp.indexable,1)=1 AND COALESCE(sp.excluded_from_audit,0)=0
""").fetchall()

def fetch(url):
    try:
        r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT, allow_redirects=True)
        if r.status_code >= 400: return (None, None)
        box = BeautifulSoup(r.content, "lxml").select_one('.collection-description-box')
        if not box: return (0, 0)
        return (len(box.get_text(' ', strip=True).split()), len(box.find_all('img')))
    except Exception:
        return (None, None)

data = {}
with ThreadPoolExecutor(max_workers=16) as ex:
    futs = {ex.submit(fetch, r[0]): r[0] for r in rows}
    for f in as_completed(futs):
        data[futs[f]] = f.result()

recs = []
for url, title, meta, h1, lvl, t1, cj_title in rows:
    handle = url.rstrip('/').split('/collections/')[-1]
    bare = SUF.sub('', title).strip()
    name = (cj_title or '').strip() or h1 or handle
    tier = (f"T{lvl}" + (f" · {t1}" if t1 else "")) if lvl else "(chưa phân tầng)"
    w, im = data.get(url, (None, None))
    tl, ml = len(bare), len(meta)
    recs.append({'tier': tier, 'lvl': int(lvl) if lvl else 9, 't1': t1 or 'zzz', 'name': name,
                 'title': bare, 'tl': tl, 'meta': meta, 'ml': ml, 'words': w, 'imgs': im, 'url': url,
                 't_bad': tl < T_LO or tl > T_HI, 'm_bad': ml < M_LO or ml > M_HI})
recs.sort(key=lambda r: (r['t1'], r['lvl'], r['name'].lower()))

wb = openpyxl.Workbook()

# ===== Sheet 1: Tổng quan =====
ws = wb.active; ws.title = "Tổng quan (211)"
H1 = ['STT', 'Phân tầng', 'Tên collection', 'Title (SEO)', 'Title\n(ký tự)',
      'Meta description', 'Meta\n(ký tự)', 'Mô tả\n(số từ)', 'Ảnh\n(mô tả)', 'Link live']
style_header(ws, H1)
for i, r in enumerate(recs, 1):
    rr = i + 1
    vals = [i, r['tier'], r['name'], r['title'], r['tl'], r['meta'], r['ml'],
            r['words'] if r['words'] is not None else 'ERR',
            r['imgs'] if r['imgs'] is not None else 'ERR', r['url']]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(rr, ci, v); cell.border = B
        cell.alignment = CEN if ci in (1, 5, 7, 8, 9) else WRAP
        if i % 2 == 0 and ci in (2, 3, 4, 6): cell.fill = ALT
    if r['tl'] and r['t_bad'] or (r['tl'] and not (48 <= r['tl'] <= 58)): ws.cell(rr, 5).fill = BAD if r['t_bad'] else WARN
    if r['ml'] and r['m_bad'] or (r['ml'] and not (140 <= r['ml'] <= 160)): ws.cell(rr, 7).fill = BAD if r['m_bad'] else WARN
    if r['words'] is not None and r['words'] < WORD_FLOOR: ws.cell(rr, 8).fill = WARN
    ws.cell(rr, 10).hyperlink = r['url']; ws.cell(rr, 10).font = Font(color="0563C1", underline="single")
for ci, w in enumerate([5, 22, 26, 46, 8, 50, 8, 8, 7, 30], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 32; ws.freeze_panes = 'C2'; ws.auto_filter.ref = f"A1:J{len(recs)+1}"

# ===== Sheet 2: Cần sửa title-meta =====
bad = [r for r in recs if r['t_bad'] or r['m_bad']]
bad.sort(key=lambda r: (not (r['t_bad'] and r['m_bad']), r['t1'], r['lvl'], r['name'].lower()))
ws2 = wb.create_sheet("Cần sửa title-meta")
H2 = ['STT', 'Phân tầng', 'Tên collection', 'Title (SEO)', 'Title\n(ký tự)',
      'Meta description', 'Meta\n(ký tự)', 'Lỗi', 'Link live']
style_header(ws2, H2, PatternFill("solid", fgColor="C0392B"))
for i, r in enumerate(bad, 1):
    rr = i + 1
    flags = []
    if r['tl'] < T_LO: flags.append(f'TITLE ngắn ({r["tl"]})')
    elif r['tl'] > T_HI: flags.append(f'TITLE dài ({r["tl"]})')
    if r['ml'] < M_LO: flags.append(f'META ngắn ({r["ml"]})')
    elif r['ml'] > M_HI: flags.append(f'META dài ({r["ml"]})')
    vals = [i, r['tier'], r['name'], r['title'], r['tl'], r['meta'], r['ml'], '; '.join(flags), r['url']]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(rr, ci, v); cell.border = B
        cell.alignment = CEN if ci in (1, 5, 7) else WRAP
        if i % 2 == 0 and ci in (2, 3, 4, 6, 8): cell.fill = ALT
    if r['t_bad']: ws2.cell(rr, 5).fill = BAD
    if r['m_bad']: ws2.cell(rr, 7).fill = BAD
    ws2.cell(rr, 9).hyperlink = r['url']; ws2.cell(rr, 9).font = Font(color="0563C1", underline="single")
for ci, w in enumerate([5, 22, 26, 46, 8, 52, 8, 26, 30], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[1].height = 30; ws2.freeze_panes = 'C2'; ws2.auto_filter.ref = f"A1:I{len(bad)+1}"

wb.save(OUT)
print(f"Saved 1 file, {len(wb.sheetnames)} sheet: {wb.sheetnames}")
print(f"  Tổng quan: {len(recs)} | Cần sửa: {len(bad)} (title {sum(1 for r in recs if r['t_bad'])} / meta {sum(1 for r in recs if r['m_bad'])})")
